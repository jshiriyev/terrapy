import calendar

from datetime import datetime
from dateutil import parser
from dateutil import relativedelta

import math
import os
import re

import numpy as np
import openpyxl as opxl
import lasio

if __name__ == "__main__":
    import setup

class Files():

    def __init__(self):

        pass

    def set_homepath(self,homepath):

        self.homepath = homepath

    def get_filename(self,path,prefix,extension,homeFlag=False):

        if homeFlag and hasattr(self,"homepath"):
            path = os.path.join(self.homepath,path)

        filenames = os.listdir(path)

        return [filename for filename in filenames if filename.startswith(prefix) and filename.endswith(extension)]

class Frame(Files):

    # MAIN DATA FRAME

    def __init__(self,headers=None,homepath=None,filepath=None,skiplines=0,headerline=None,comment=None,endline=None,endfile=None,**kwargs):

        # There are two uses of dataset, case 1 or case 2:
        #   case 1: headers
        #   case 2: filepath,skiplines,headerline
        #    - reading plain text: comment,endline,endfile
        #    - reading scpecial extensions: **kwargs
        #   case 3: no input
        
        if headers is not None:
            self._headers = headers
            self._running = [np.array([])]*len(self._headers)

        elif filepath is not None:
            self.filepath = filepath
            self.skiplines = skiplines

            if headerline is None:
                self.headerline = skiplines-1
            elif headerline<skiplines:
                self.headerline = headerline
            else:
                self.headerline = skiplines-1

            self.comment = comment
            self.endline = endline
            self.endfile = endfile

            self.filename = os.path.split(self.filepath)[1]
            self.extension = os.path.splitext(self.filepath)[1]

            if any([self.extension==extension for extension in self.special_extensions]):
                self.read_special(**kwargs)
            else:
                self.read()

        else:
            self._headers = []
            self._running = []

        self.headers = self._headers
        self.running = [np.asarray(column) for column in self._running]

    def read(self):

        # While looping inside the file it does not read lines:
        # - starting with comment phrase, e.g., comment = "--"
        # - after the end of line phrase, e.g., endline = "/"
        # - after the end of file keyword e.g., endfile = "END"

        self._running = []

        with open(self.filepath,"r") as text:

            for line in text:

                line = line.split('\n')[0].strip()

                if self.endline is not None:
                    line = line.strip(self.endline)

                line = line.strip()
                line = line.strip("\t")
                line = line.strip()

                if line=="":
                    continue

                if self.comment is not None:
                    if line[:len(self.comment)] == self.comment:
                        continue

                if self.endfile is not None:
                    if line[:len(self.endfile)] == self.endfile:
                        break

                self._running.append([line])

        self.title = []

        for _ in range(self.skiplines):
            self.title.append(self._running.pop(0))

        num_cols = len(self._running[0])

        if self.skiplines==0:
            self._headers = ["Column #"+str(index) for index in range(num_cols)]
        elif self.skiplines!=0:
            self._headers = self.title[self.headerline]

        nparray = np.array(self._running).T

        self._running = [np.asarray(column) for column in nparray]

    def set_header(self,header_index,header):

        self._headers[header_index] = header

        self.headers = self._headers

    def texttocolumn(self,header_index=None,header=None,deliminator=None,maxsplit=None):

        if header_index is None:
            header_index = self._headers.index(header)

        header_string = self._headers[header_index]
        # header_string = re.sub(deliminator+'+',deliminator,header_string)

        headers = header_string.split(deliminator)

        if maxsplit is not None:
            for index in range(maxsplit):
                if len(headers)<index+1:
                    headers.append("col ##"+string(header_index+index))

        column_to_split = np.asarray(self._running[header_index])

        running = []

        for index,string in enumerate(column_to_split):

            # string = re.sub(deliminator+'+',deliminator,string)
            row = np.char.split(string,deliminator).tolist()

            if maxsplit is not None:
                while len(row)<maxsplit:
                    row.append("")

            running.append(row)

        running = np.array(running,dtype=str).T

        self._headers.pop(header_index)
        self._running.pop(header_index)

        for header,column in zip(headers,running):
            self._headers.insert(header_index,header)
            self._running.insert(header_index,column)
            header_index += 1

        # line = re.sub(r"[^\w]","",line)
        # line = "_"+line if line[0].isnumeric() else line
        # vmatch = np.vectorize(lambda x: bool(re.compile('[Ab]').match(x)))
        
        self.headers = self._headers
        self.running = [np.asarray(column) for column in self._running]

    def columntotext(self,header_new,header_indices=None,headers=None,string=None):

        if header_indices is None:
            header_indices = [self._headers.index(header) for header in headers]

        if string is None:
            string = ("{} "*len(header_indices)).strip()

        vprint = np.vectorize(lambda *args: string.format(*args))

        column_new = [np.asarray(self._running[index]) for index in header_indices]

        column_new = vprint(*column_new)

        self._headers.append(header_new)
        self._running.append(column_new)

        self.headers = self._headers
        self.running = [np.asarray(column) for column in self._running]

    def astype(self,header_index=None,header=None,dtype=None,nonetozero=False,datestring=False,shiftmonths=0):

        if header_index is None:
            header_index = self._headers.index(header)

        if datestring:

            def shifting(string):
                date = parser.parse(string)+relativedelta.relativedelta(months=shiftmonths)
                days = calendar.monthrange(date.year,date.month)[1]
                return datetime(date.year,date.month,days)

            if dtype is None:
                if shiftmonths != 0:
                    vdate = np.vectorize(lambda x: shifting(x))
                else:
                    vdate = np.vectorize(lambda x: parser.parse(x))
            else:
                if shiftmonths != 0:
                    vdate = np.vectorize(lambda x: dtype(shifting(x)))
                else:
                    vdate = np.vectorize(lambda x: dtype(parser.parse(x)))
            
        else:

            if nonetozero:
                vdate = np.vectorize(lambda x: dtype(x) if x is not None else dtype(0))
            else:
                vdate = np.vectorize(lambda x: dtype(x))
            
        self._running[header_index] = vdate(self._running[header_index])

        self.running[header_index] = np.asarray(self._running[header_index])

    def edit_strcolumn(self,fstring,header_index=None,header=None):

        if header_index is None:
            header_index = self._headers.index(header)

        editor = np.vectorize(lambda x: fstring.format(x))

        self._running[header_index] = editor(self._running[header_index])

        self.running[header_index] = np.asarray(self._running[header_index])

    def upper(self,header_index=None,header=None):

        if header_index is None:
            header_index = self._headers.index(header)

        self._running[header_index] = np.char.upper(self._running[header_index])

    def set_columns(self,columns,header_indices=None,headers=None):

        if headers is None:
            if header_indices is None:
                headers = ["Col ##{}".format(i) for i in range(len(columns))]
        
        if header_indices is None:
            for header,column in zip(headers,columns):
                self._headers.append(header)
                self._running.append(column)
        else:
            for index,column in zip(header_indices,columns):
                self._running[index] = column

        self.headers = self._headers
        self.running = [np.asarray(column) for column in self._running]

    def set_rows(self,rows,row_indices=None):
        
        for row in rows:

            if row_indices is None:
                for col_index,column in enumerate(self._running):
                    self._running[col_index] = np.append(column,row[col_index])
            else:
                for col_index, _ in enumerate(self._running):
                    self._running[col_index][row_indices] = row[col_index]

            self.running = [np.asarray(column) for column in self._running]

    def get_rows(self,row_indices=None,match=None):

        if row_indices is None:
            if match is None:
                row_indices = range(self._running[0].size)
            else:
                column_index,phrase = match
                conditional = self._running[column_index]==phrase
                row_indices = np.arange(self._running[0].size)[conditional]

        elif type(row_indices)==int:
            row_indices = [row_indices]

        rows = [[column[index] for column in self._running] for index in row_indices]
        
        return rows

    def get_columns(self,header_indices=None,headers=None,match=None,inplace=False,returnFlag=False):

        if header_indices is None:
            header_indices = [self._headers.index(header) for header in headers]

        if match is None:
            conditional = np.full(self._running[0].shape,True)
        else:
            conditional = self._running[match[0]]==match[1]

        if inplace:
            self._headers = [self._headers[index] for index in header_indices]
            self._running = [self._running[index][conditional] for index in header_indices]

            self.headers = self._headers
            self.running = [np.asarray(column) for column in self._running]

        else:
            if not returnFlag:
                self.headers = [self._headers[index] for index in header_indices]
                self.running = [np.asarray(self._running[index][conditional]) for index in header_indices]
            else:
                return [self._running[index][conditional] for index in header_indices]

    def del_rows(self,row_indices=None,noneColIndex=None,inplace=False):

        all_rows = np.array([np.arange(self._running[0].size)])

        if row_indices is None:
            row_indices = [index for index,val in enumerate(self.running[noneColIndex]) if val is None]
        
        row_indices = np.array(row_indices).reshape((-1,1))

        comp_mat = all_rows==row_indices

        keep_index = ~np.any(comp_mat,axis=0)

        if inplace:
            self._running = [column[keep_index] for column in self._running]
            self.running = [np.asarray(column) for column in self._running]
        else:
            self.running = [np.asarray(column[keep_index]) for column in self._running]

    def fill_nones(self,col_indices=None,inplace=False):

        if col_indices is None:
            col_indices = range(len(self._running))

        for col_index in col_indices:

            column = self._running[col_index]

            row_indices = [row_index for row_index,val in enumerate(column) if val is None]

            for row_index in row_indices:
                column[row_index] = column[row_index-1]

            if inplace:
                self._running[col_index] = column
                self.running[col_index] = np.asarray(column)
            else:
                self.running[col_index] = np.asarray(column)
            
    def sort(self,header_indices=None,headers=None,reverse=False,inplace=False,returnFlag=False):

        if header_indices is None:
            header_indices = [self.headers.index(header) for header in headers]

        columns = [self._running[index] for index in header_indices]

        columns.reverse()

        sort_index = np.lexsort(columns)

        if reverse:
            sort_index = np.flip(sort_index)

        if inplace:
            self._running = [column[sort_index] for column in self._running]
            self.running = [np.asarray(column) for column in self._running]
        else:
            self.running = [np.asarray(column[sort_index]) for column in self._running]

        if returnFlag:
            return sort_index

    def filter(self,header_index=None,header=None,keywords=None,regex=None,inplace=False):

        if header_index is None:
            header_index = self._headers.index(header)

        if keywords is not None:
            match_array = np.array(keywords).reshape((-1,1))
            match_index = np.any(self._running[header_index]==match_array,axis=0)
        else:
            match_vectr = np.vectorize(lambda x:bool(re.compile(regex).match(x)))
            match_index = match_vectr(self._running[header_index])

        if inplace:
            self._running = [column[match_index] for column in self._running]
            self.running = [np.asarray(column) for column in self._running]
        else:
            self.running = [np.asarray(column[match_index]) for column in self._running]

    def filter_invert(self):

        self.running = [np.asarray(column) for column in self._running]

    def write(self,filepath,fstring=None,**kwargs):

        header_fstring = ("{}\t"*len(self._headers))[:-1]+"\n"

        if fstring is None:
            running_fstring = ("{}\t"*len(self._headers))[:-1]+"\n"
        else:
            running_fstring = fstring

        vprint = np.vectorize(lambda *args: running_fstring.format(*args))

        with open(filepath,"w",encoding='utf-8') as wfile:
            wfile.write(header_fstring.format(*self._headers))
            for line in vprint(*self._running):
                wfile.write(line)

class Excel(Frame):

    def __init__(self,filepaths=None,headers=None):

        self.files = []

        if filepaths is not None:
            for filepath in filepaths:
                self.add_file(filepath)

        if headers is not None:
            self._headers = headers
            self._running = [np.array([])]*len(self._headers)
        else:
            self._headers = []
            self._running = []

        self.headers = self._headers
        self.running = [np.asarray(column) for column in self._running]

    def add_file(self,filepath,homeFlag=False):

        if homeFlag and hasattr(self,"homepath"):

            filepath = os.path.join(self.homepath,filepath)

        self.files.append(opxl.load_workbook(filepath,read_only=True,data_only=True))

    def get_sheetname(self,keyword):

        # This method returns sheetname similar to the keyword

        for sheetname in self.files[0].sheetnames:
            if sheetname[:len(keyword)]==keyword:
                return sheetname
            else:
                return keyword

    def read_headers(self,sheetname,row_index):

        row = self.files[0][sheetname].iter_rows(min_row=row_index,min_col=1,
            max_row=row_index,max_col=None,values_only=True)

        headers = list(row)[0]

        [self._headers.append(re.sub(r"[^\w]","",header)) for header in headers]

        self.headers = self._headers

    def read(self,sheetname,min_row=1,min_col=1,max_row=None,max_col=None):

        rows = self.files[0][sheetname].iter_rows(min_row=min_row,min_col=min_col,
            max_row=max_row,max_col=max_col,values_only=True)

        rows = list(rows)

        if len(self._headers)==0:

            [self._headers.append("Col {}".format(i)) for i in range(len(rows[0]))]

            self.headers = self._headers

        if len(self._running)==0:
            self._running = [np.array([])]*len(rows[0])

        self.set_rows(rows)

    def write(self):

        wb = opxl.Workbook()

        sheet = wb.active

        if sheet_title is not None:
            sheet.title = sheet_title

        for line in running:
            sheet.append(line)

        wb.save(filepath)

    def close(self,index=None):

        if index is None:
            [file._archive.close() for file in self.files]
        else:
            self.files[index]._archive.close()

class VTKit(Files):

    def __init__(self):

        pass

    def read(self,):

        pass

    def write(self,):

        pass

class History(Frame):

    # KEYWORDS: DATES,COMPDATMD,COMPORD,WCONHIST,WCONINJH,WEFAC,WELOPEN 

    headers    = ["DATE","KEYWORD","DETAILS",]

    dates      = " {} / "#.format(date)
    welspecs   = " '{}'\t1*\t2* / "
    compdatop  = " '{}'\t1*\t{}\t{}\tMD\t{}\t2*\t0.14 / "#.format(wellname,top,bottom,optype)
    compdatsh  = " '{}'\t1*\t{}\t{}\tMD\t{} / "#.format(wellname,top,bottom,optype)
    compord    = " '{}'\tINPUT\t/ "#.format(wellname)
    prodhist   = " '{}'\tOPEN\tORAT\t{}\t{}\t{} / "#.format(wellname,oilrate,waterrate,gasrate)
    injhist    = " '{}'\tWATER\tOPEN\t{}\t7*\tRATE / "#.format(wellname,waterrate)
    wefac      = " '{}'\t{} / "#.format(wellname,efficiency)
    welopen    = " '{}'\tSHUT\t3* / "#.format(wellname)

    def __init__(self):

        pass

    def read(self):

        pass

    def set_subheaders(self,header_index=None,header=None,regex=None,regex_builtin="INC_HEADERS",title="SUB-HEADERS"):

        nparray = np.array(self._running[header_index])

        if regex is None and regex_builtin=="INC_HEADERS":
            regex = r'^[A-Z]+$'                         #for strings with only capital letters no digits
        elif regex is None and regex_builtin=="INC_DATES":
            regex = r'^\d{1,2} [A-Za-z]{3} \d{2}\d{2}?$'   #for strings with [1 or 2 digits][space][3 capital letters][space][2 or 4 digits], e.g. DATES

        vmatch = np.vectorize(lambda x: bool(re.compile(regex).match(x)))

        match_index = vmatch(nparray)

        firstocc = np.argmax(match_index)

        lower = np.where(match_index)[0]
        upper = np.append(lower[1:],nparray.size)

        repeat_count = upper-lower-1

        match_content = nparray[match_index]

        nparray[firstocc:][~match_index[firstocc:]] = np.repeat(match_content,repeat_count)

        self._headers.insert(header_index,title)
        self._running.insert(header_index,np.asarray(nparray))

        for index,column in enumerate(self._running):
            self._running[index] = np.array(self._running[index][firstocc:][~match_index[firstocc:]])

        self.headers = self._headers
        self.running = [np.asarray(column) for column in self._running]

    def get_wells(self,wellname=None):

        pass

    def write(self):

        path = os.path.join(self.workdir,self.schedule_filename)

        with open(path,"w",encoding='utf-8') as wfile:

            welspec = schedule.running[1]=="WELSPECS"
            compdat = schedule.running[1]=="COMPDATMD"
            compord = schedule.running[1]=="COMPORD"
            prodhst = schedule.running[1]=="WCONHIST"
            injdhst = schedule.running[1]=="WCONINJH"
            wefffac = schedule.running[1]=="WEFAC"
            welopen = schedule.running[1]=="WELOPEN"

            for date in np.unique(schedule.running[0]):

                currentdate = schedule.running[0]==date

                currentcont = schedule.running[1][currentdate]

                wfile.write("\n\n")
                wfile.write("DATES\n")
                wfile.write(self.schedule_dates.format(date.strftime("%d %b %Y").upper()))
                wfile.write("\n")
                wfile.write("/\n\n")

                if any(currentcont=="WELSPECS"):
                    indices = np.logical_and(currentdate,welspec)
                    wfile.write("WELSPECS\n")
                    for detail in schedule.running[2][indices]:
                        wfile.write(detail)
                        wfile.write("\n")
                    wfile.write("/\n\n")

                if any(currentcont=="COMPDATMD"):
                    indices = np.logical_and(currentdate,compdat)
                    wfile.write("COMPDATMD\n")
                    for detail in schedule.running[2][indices]:
                        wfile.write(detail)
                        wfile.write("\n")
                    wfile.write("/\n\n")

                if any(currentcont=="COMPORD"):
                    indices = np.logical_and(currentdate,compord)
                    wfile.write("COMPORD\n")
                    for detail in schedule.running[2][indices]:
                        wfile.write(detail)
                        wfile.write("\n")
                    wfile.write("/\n\n")

                if any(currentcont=="WCONHIST"):
                    indices = np.logical_and(currentdate,prodhst)
                    wfile.write("WCONHIST\n")
                    for detail in schedule.running[2][indices]:
                        wfile.write(detail)
                        wfile.write("\n")
                    wfile.write("/\n\n")

                if any(currentcont=="WCONINJH"):
                    indices = np.logical_and(currentdate,injdhst)
                    wfile.write("WCONINJH\n")
                    for detail in schedule.running[2][indices]:
                        wfile.write(detail)
                        wfile.write("\n")
                    wfile.write("/\n\n")

                if any(currentcont=="WEFAC"):
                    indices = np.logical_and(currentdate,wefffac)
                    wfile.write("WEFAC\n")
                    for detail in schedule.running[2][indices]:
                        wfile.write(detail)
                        wfile.write("\n")
                    wfile.write("/\n\n")

                if any(currentcont=="WELOPEN"):
                    indices = np.logical_and(currentdate,welopen)
                    wfile.write("WELOPEN\n")
                    for detail in schedule.running[2][indices]:
                        wfile.write(detail)
                        wfile.write("\n")
                    wfile.write("/\n\n")

class LogASCII(Files):

    def __init__(self,filepaths=None):

        self.files = []

        self.headers = []

        if filepaths is not None:
            for filepath in filepaths:
                self.add_file(filepath)

    def add_file(self,filepath):

        las = lasio.read(filepath)

        self.files.append(las)

        self.headers.append(las.keys())

    def print_well_info(self,index=None):

        if index is not None:
            print("\n\tWELL #{}".format(self.files[index].well.WELL.value))
            for item in self.files[index].sections["Well"]:
                print(f"{item.descr} ({item.mnemonic}):\t\t{item.value}")
        else:
            for las in self.files:
                print("\n\tWELL #{}".format(las.well.WELL.value))
                for item in las.sections["Well"]:
                    print(f"{item.descr} ({item.mnemonic}):\t\t{item.value}")

    def print_curve_info(self,index=None,mnemonic_space=33,tab_space=8):

        with open('terrapy_LogASCII.out','w') as file:

            def print_func(index):
                las = self.files[index]
                file.write("\n\tLOG NUMBER {}\n".format(index))
                # print("\n\tLOG NUMBER {}".format(index))
                for count,curve in enumerate(las.curves):
                    minXval = np.nanmin(curve.data)
                    maxXval = np.nanmax(curve.data)
                    tab_num = math.ceil((mnemonic_space-len(curve.mnemonic))/tab_space)
                    tab_spc = "\t"*tab_num if tab_num>0 else "\t"
                    file.write("Curve: {}{}Units: {}\tMin: {}\tMax: {}\tDescription: {}\n".format(
                        curve.mnemonic,tab_spc,curve.unit,minXval,maxXval,curve.descr))
                    # print("Curve: {}{}Units: {}\tMin: {}\tMax: {}\tDescription: {}".format(
                    #     curve.mnemonic,tab_spc,curve.unit,minXval,maxXval,curve.descr))

            if index is not None:
                print_func(index)
            else:
                [print_func(index) for index in range(len(self.files))]

    def flip(self,fileID):

        for index,curve in enumerate(self.files[fileID].curves):

            self.files[fileID].curves[index].data = np.flip(curve.data)
            
    def set_interval(self,top,bottom,fileID=None):

        if fileID is None:
            fileIDs = range(len(self.files))
        else:
            fileIDs = range(fileID,fileID+1)

        self.top = top
        self.bottom = bottom

        self.gross_thickness = self.bottom-self.top

        for indexI in fileIDs:

            las = self.files[indexI]

            try:
                depth = las["MD"]
            except KeyError:
                depth = las["DEPT"]

            depth_cond = np.logical_and(depth>self.top,depth<self.bottom)

            for curveID,curve in enumerate(las.curves):
                self.files[indexI].curves[curveID].data = curve.data[depth_cond]

    def get_interval(self,top,bottom,fileID=None,curveID=None):

        returningList = []

        if fileID is None:
            fileIDs = range(len(self.files))
        else:
            fileIDs = range(fileID,fileID+1)

        for indexI in fileIDs:

            las = self.files[indexI]

            try:
                depth = las["MD"]
            except KeyError:
                depth = las["DEPT"]

            depth_cond = np.logical_and(depth>top,depth<bottom)

            if curveID is None:
                returningList.append(depth_cond)
            else:
                returningList.append(las.curves[curveID].data[depth_cond])

        return returningList

    def resample(self,depthsFID=None,depths=None,fileID=None,curveID=None):

        """

        depthsFID:  The index of file id from which to take new depths
                    where new curve data will be calculated;

        depths:     The numpy array of new depths
                    where new curve data will be calculated;
        
        fileID:     The index of file to resample;
                    If None, all files will be resampled;
        
        curveID:    The index of curve in the las file to resample;
                    If None, all curves in the file will be resampled;
                    Else if fileID is not None, resampled data will be returned;

        """

        if depthsFID is not None:
            try:
                depths = self.files[depthsFID]["MD"]
            except KeyError:
                depths = self.files[depthsFID]["DEPT"]

        if fileID is None:
            fileIDs = range(len(self.files))
        else:
            fileIDs = range(fileID,fileID+1)

        for indexI in fileIDs:

            if depthsFID is not None:
                if indexI==depthsFID:
                    continue

            las = self.files[indexI]

            try:
                depths_original = las["MD"]
            except KeyError:
                depths_original = las["DEPT"]

            lowerend = depths<depths_original.min()
            upperend = depths>depths_original.max()

            interior = np.logical_and(~lowerend,~upperend)

            depths_interior = depths[interior]

            diff = depths_original-depths_interior.reshape((-1,1))

            indices_lower = np.where(diff<0,diff,-np.inf).argmax(axis=1)
            indices_upper = np.where(diff>0,diff,np.inf).argmin(axis=1)

            grads = (depths_interior-depths_original[indices_lower])/(depths_original[indices_upper]-depths_original[indices_lower])

            if curveID is None:
                las.curves[0].data = depths

            if curveID is None:
                curveIDs = range(1,len(las.curves))
            else:
                curveIDs = range(curveID,curveID+1)

            for indexJ in curveIDs:

                curve = las.curves[indexJ]

                data_resampled = np.empty(depths.shape,dtype=float)

                data_resampled[lowerend] = np.nan
                data_resampled[interior] = curve.data[indices_lower]+grads*(curve.data[indices_upper]-curve.data[indices_lower])
                data_resampled[upperend] = np.nan

                if curveID is None:
                    self.files[indexI].curves[indexJ].data = data_resampled
                elif fileID is not None:
                    return data_resampled

    def merge(self,fileIDs,curveNames):

        if type(fileIDs)==int:

            try:
                depth = self.files[fileIDs]["MD"]
            except KeyError:
                depth = self.files[fileIDs]["DEPT"]

            xvals1 = self.files[fileIDs][curveNames[0]]

            for curveName in curveNames[1:]:

                xvals2 = self.files[fileIDs][curveName]

                xvals1[np.isnan(xvals1)] = xvals2[np.isnan(xvals1)]

            return depth,xvals1

        elif np.unique(np.array(fileIDs)).size==len(fileIDs):

            if type(curveNames)==str:
                curveNames = (curveNames,)*len(fileIDs)

            depth = np.array([])
            xvals = np.array([])

            for (fileID,curveName) in zip(fileIDs,curveNames):

                try:
                    depth_loc = self.files[fileID]["MD"]
                except KeyError:
                    depth_loc = self.files[fileID]["DEPT"]

                xvals_loc = self.files[fileID][curveName]

                depth_loc = depth_loc[~np.isnan(xvals_loc)]
                xvals_loc = xvals_loc[~np.isnan(xvals_loc)]

                depth_max = 0 if depth.size==0 else depth.max()

                depth = np.append(depth,depth_loc[depth_loc>depth_max])
                xvals = np.append(xvals,xvals_loc[depth_loc>depth_max])

            return depth,xvals

    def write(self,filepath,mnemonics,data,fileID=None,units=None,descriptions=None,values=None):

        """
        
        filepath:       It will write a lasio.LASFile to the given filepath
        fileID:         The file index which to write to the given filepath
                        If fileID is None, new lasio.LASFile will be created

        kwargs:         These are mnemonics, data, units, descriptions, values

        """

        if fileID is not None:

            lasfile = self.files[fileID]

        else:

            lasfile = lasio.LASFile()

            lasfile.well.DATE = datetime.today().strftime('%Y-%m-%d %H:%M:%S')

            depthExistFlag = False

            for mnemonic in mnemonics:
                if mnemonic=="MD" or mnemonic=="DEPT":
                    depthExistFlag = True
                    break

            if not depthExistFlag:
                curve = lasio.CurveItem(
                    mnemonic="DEPT",
                    unit="",
                    value="",
                    descr="Depth index",
                    data=np.arange(data[0].size))
                lasfile.append_curve_item(curve)            

        for index,(mnemonic,datum) in enumerate(zip(mnemonics,data)):

            if units is not None:
                unit = units[index]
            else:
                unit = ""

            if descriptions is not None:
                description = descriptions[index]
            else:
                description = ""

            if values is not None:
                value = values[index]
            else:
                value = ""

            curve = lasio.CurveItem(mnemonic=mnemonic,data=datum,unit=unit,descr=description,value=value)

            lasfile.append_curve_item(curve)

        with open(filepath, mode='w') as filePathToWrite:
            lasfile.write(filePathToWrite)

def resample(depths,data,depthsR):

    lowerend = depthsR<depths.min()
    upperend = depthsR>depths.max()

    interior = np.logical_and(~lowerend,~upperend)

    depths_interior = depthsR[interior]

    indices_lower = np.empty(depths_interior.shape,dtype=int)
    indices_upper = np.empty(depths_interior.shape,dtype=int)

    for index,depth in enumerate(depths_interior):

        diff = depths-depth

        indices_lower[index] = np.where(diff<0,diff,-np.inf).argmax()
        indices_upper[index] = np.where(diff>0,diff,np.inf).argmin()

    grads = (depths_interior-depths[indices_lower])/(depths[indices_upper]-depths[indices_lower])

    dataR = np.empty(depthsR.shape,dtype=float)

    dataR[lowerend] = np.nan
    dataR[interior] = data[indices_lower]+grads*(data[indices_upper]-data[indices_lower])
    dataR[upperend] = np.nan

    return dataR

def cyrilictolatin(string):

    aze_cyril_lower = [
        "а","б","ж","ч","д",
        "е","я","ф","э","ь",
        "щ","х","ы","и","ъ",
        "к","г","л","м","н",
        "о","ю","п","р","с",
        "ш","т","у","ц","в",
        "й","з"]

    aze_latin_lower = [
        "a","b","c","ç","d",
        "e","ə","f","g","ğ",
        "h","x","ı","i","j",
        "k","q","l","m","n",
        "o","ö","p","r","s",
        "ş","t","u","ü","v",
        "y","z"]

    aze_cyril_upper = [
        "А","Б","Ҹ","Ч","Д",
        "Е","Я","Ф","Ҝ","Ғ",
        "Щ","Х","Ы","И","Ъ",
        "К","Г","Л","М","Н",
        "О","Ю","П","Р","С",
        "Ш","Т","У","Ц","В",
        "Й","З"]

    aze_latin_upper = [
        "A","B","C","Ç","D",
        "E","Ə","F","G","Ğ",
        "H","X","I","İ","J",
        "K","Q","L","M","N",
        "O","Ö","P","R","S",
        "Ş","T","U","Ü","V",
        "Y","Z"]

    for cyril,latin in zip(aze_cyril_lower,aze_latin_lower):
        string.replace(cyril,latin)

    for cyril,latin in zip(aze_cyril_upper,aze_latin_upper):
        string.replace(cyril,latin)

    return string

if __name__ == "__main__":

    import stream.tests