"""

The task 

"""

import datetime

import matplotlib.pyplot as plt

import numpy as np

import openpyxl

inputfile = 'C:\\Users\\javid.s\\Documents\\South_GD_history_match_06_July_2020.xlsx'

##wells = ["409","412","413","414","415","417","418","430","field"]

class read_production_data():

    def __init__(self,inputfile):

        self.inputfile = inputfile
        
        self.wb = openpyxl.load_workbook(self.inputfile)

    def get_well(self,wellname):

        self.ws = self.wb[wellname]

        Date = list(self.ws.iter_cols(min_row=2,max_col=1,values_only=True))[0]

        Head = np.array(list(self.ws.iter_rows(max_row=1,min_col=2,values_only=True))[0])
        Body = np.array(list(self.ws.iter_cols(min_row=2,min_col=2,values_only=True)))

        cols_calc = (0,2,4,6,8,10,12)
        cols_hist = (1,3,5,7,9,11,13)

        Head = Head[cols_calc,]
        Calc = Body[cols_calc,]
        Hist = Body[cols_hist,]

        head = []

        for i,h in enumerate(Head):
            string = h.split(":")[1].split(",")[0].strip().replace(" ","_")
            head.append(string)

        self.date = Date
        self.head = head
        self.calc = Calc
        self.hist = Hist

    def get_field(self,fieldname):

        self.ws = self.wb[fieldname]

        Date = list(self.ws.iter_cols(min_row=2,max_col=1,values_only=True))[0]

        Head = np.array(list(self.ws.iter_rows(max_row=1,min_col=2,values_only=True))[0])
        Body = np.array(list(self.ws.iter_cols(min_row=2,min_col=2,values_only=True)))

        cols_calc = (0,2,4,6,8,10,12)
        cols_hist = (1,3,5,7,9,11)

        Head = Head[cols_calc,]
        Calc = Body[cols_calc,]
        Hist = Body[cols_hist,]

        head = []

        for i,h in enumerate(Head):
            string = h.split(":")[1].split(",")[0].strip().replace(" ","_")
            head.append(string)

        self.date = Date
        self.head = head
        self.calc = Calc
        self.hist = Hist

P = read_production_data(inputfile)

P.get_field("field")

fig,axs = plt.subplots(2,2,figsize=(14,8))

axs0 = axs[0,0].twinx()
axs1 = axs[0,1].twinx()
axs2 = axs[1,0].twinx()

axs[0,0].plot(P.date,P.calc[1],c='k')
axs[0,1].plot(P.date,P.calc[3]/1000,c='k')
axs[1,0].plot(P.date,P.calc[5],c='k')
axs[1,1].plot(P.date,P.calc[6],c='k')

axs0.plot(P.date,P.calc[0]/1000,c='k')
axs1.plot(P.date,P.calc[2]/1000000,c='k')
axs2.plot(P.date,P.calc[4]/1000,c='k')

axs[0,0].plot(P.date,P.hist[1],'--',c='g')
axs[0,1].plot(P.date,P.hist[3]/1000,'--',c='r')
axs[1,0].plot(P.date,P.hist[5],'--',c='b')
##axs[1,1].plot(P.date,P.hist[6],'--',c='m')

axs0.plot(P.date,P.hist[0]/1000,'--',c='g')
axs1.plot(P.date,P.hist[2]/1000000,'--',c='r')
axs2.plot(P.date,P.hist[4]/1000,'--',c='b')

axs[0,0].set_xlabel("Date")
axs[0,1].set_xlabel("Date")
axs[1,0].set_xlabel("Date")
axs[1,1].set_xlabel("Date")

axs[0,0].set_ylabel("Liquid Rate, sm3/day")
axs[0,1].set_ylabel("Gas Rate, th. sm3/day")
axs[1,0].set_ylabel("Liquid Rate, sm3/day")
axs[1,1].set_ylabel("Pressure, Bars")

axs0.set_ylabel("Liquid Volume, th. sm3")
axs1.set_ylabel("Surface Gas Volume, mln. sm3")
axs2.set_ylabel("Liquid Volume, th. sm3")

axs[0,0].grid()
axs[0,1].grid()
axs[1,0].grid()
axs[1,1].grid()

plt.tight_layout()
plt.show()
                
