"""
The task is about the history matches. When we do changes to input files and run
it, we want to compare two sets of production-history data. It is kinda
having set point between each runs that allow us to select best path to follow.
"""

import datetime
import os

import matplotlib.pyplot as plt

import numpy as np

import openpyxl

import tkinter as tk

from tkinter import filedialog

from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

class plot_production():

    def __init__(self,window):

        self.root = window

        self.initUI()

    def initUI(self):

        self.root.title("BEOC-Subsurface")
        
        self.root.configure(background="white")

        menubar = tk.Menu(self.root)
        
        self.root.config(menu=menubar)

        fileMenu = tk.Menu(menubar,tearoff="Off")

        fileMenu.add_command(label="Open",command=self.select_path)

        menubar.add_cascade(label="File",menu=fileMenu)

        self.frame = tk.Frame(self.root,width=300,height=200)

        tk.Grid.rowconfigure(self.frame,0,weight=1)

        tk.Grid.columnconfigure(self.frame,0,weight=1)
        tk.Grid.columnconfigure(self.frame,1,weight=1)

        self.frame.configure(background="white")

        self.listbox = tk.Listbox(self.frame,width=10,height=30)
        self.listbox.grid(row=0,column=0,sticky=tk.NSEW)

        self.listbox.bind('<Double-1>',self.select_list_entry)

        self.button = tk.Button(self.frame,text="Plot Graph",command=self.select_sheet)
        self.button.grid(row=1,column=0)

        self.plot_frame = tk.Frame(self.frame,width=250,height=30)
        self.plot_frame.grid(row=0,column=1,sticky=tk.NSEW)

        self.plot_frame.canvas_widget = None

        self.status = tk.Listbox(self.frame,width=250,height=5)
        self.status.grid(row=1,column=1,sticky=tk.EW)

        self.frame.pack(side=tk.LEFT,expand=1,fill=tk.BOTH)

    def select_path(self):

        self.filepath = filedialog.askopenfilename(
            title = "Select a File",
            initialdir = os.getcwd(),
            filetypes = (("Excel files","*.xl*"),("All files","*.*")))

        if not self.filepath:
            return

        self.inputfile = self.filepath
        
        self.wb = openpyxl.load_workbook(self.inputfile)

        status = "Imported \""+self.filepath+"\"."

        for sheet in self.wb.sheetnames:
            self.listbox.insert(tk.END,sheet)
        
        self.status.insert(tk.END,status)
        self.status.see(tk.END)

    def select_list_entry(self,event):

        self.select_sheet()

    def select_sheet(self):

        if not self.listbox.curselection():
            return

        if self.plot_frame.canvas_widget:
            self.plot_frame.canvas_widget.destroy()
            
        sheetname = self.listbox.get(self.listbox.curselection())

        self.ws = self.wb[sheetname]

        Head = list(self.ws.iter_rows(max_row=1,min_col=2,values_only=True))[0]
        Body = list(self.ws.iter_cols(min_row=2,min_col=2,values_only=True))

        fig,axs = plt.subplots(2,2,tight_layout=True)

        self.plot_frame.canvas = FigureCanvasTkAgg(fig,self.plot_frame)

        self.plot_frame.date = list(self.ws.iter_cols(min_row=2,max_col=1,values_only=True))[0]

        for i,head in enumerate(Head):
            string = head.split(":")[1].split(",")[0].strip().replace(" ","_")
            string = string.replace("(","").replace(")","").replace(".","")
            setattr(self.plot_frame,string,np.array(Body[i]))

        self.plot_frame.canvas_widget = self.plot_frame.canvas.get_tk_widget()
        
        self.plot_frame.canvas_widget.pack(fill=tk.BOTH,expand=1)
        
        try: axs[0,0].plot(self.plot_frame.date,self.plot_frame.Oil_Rate,c='k')
        except: pass
        try: axs[0,0].plot(self.plot_frame.date,self.plot_frame.Oil_Rate_H,'--',c='g')
        except: pass
        axs0 = axs[0,0].twinx()
        try: axs0.plot(self.plot_frame.date,self.plot_frame.Oil_Total/1000,c='k')
        except: pass
        try: axs0.plot(self.plot_frame.date,self.plot_frame.Oil_Total_H/1000,'--',c='g')
        except: pass
        
        try: axs[0,1].plot(self.plot_frame.date,self.plot_frame.Gas_Rate/1000,c='k')
        except: pass
        try: axs[0,1].plot(self.plot_frame.date,self.plot_frame.Gas_Rate_H/1000,'--',c='r')
        except: pass
        axs1 = axs[0,1].twinx()
        try: axs1.plot(self.plot_frame.date,self.plot_frame.Gas_Total/1000000,c='k')
        except: pass
        try: axs1.plot(self.plot_frame.date,self.plot_frame.Gas_Total_H/1000000,'--',c='r')
        except: pass
        
        try: axs[1,0].plot(self.plot_frame.date,self.plot_frame.Water_Rate,c='k')
        except: pass
        try: axs[1,0].plot(self.plot_frame.date,self.plot_frame.Water_Rate_H,'--',c='b')
        except: pass
        axs2 = axs[1,0].twinx()
        try: axs2.plot(self.plot_frame.date,self.plot_frame.Water_Total/1000,c='k')
        except: pass
        try: axs2.plot(self.plot_frame.date,self.plot_frame.Water_Total_H/1000,'--',c='b')
        except: pass
        
        try: axs[1,1].plot(self.plot_frame.date,self.plot_frame.Bottom_Hole_Pressure,c='k')
        except: axs[1,1].plot(self.plot_frame.date,self.plot_frame.Avg_Pressure,c='k')
        try: axs[1,1].plot(self.plot_frame.date,self.plot_frame.Bottom_Hole_Pressure_H,'--',c='m')
        except: pass
        
        axs[0,0].set_xlabel("Date")
        axs[0,1].set_xlabel("Date")
        axs[1,0].set_xlabel("Date")
        axs[1,1].set_xlabel("Date")

        axs[0,0].set_ylabel("Liquid Rate, sm3/day")
        axs[0,1].set_ylabel("Gas Rate, th. sm3/day")
        axs[1,0].set_ylabel("Liquid Rate, sm3/day")
        axs[1,1].set_ylabel("Pressure, Bars")

        axs0.set_ylabel("Liquid Volume, th. sm3")
        axs1.set_ylabel("Surface Gas Volume, mln. sm3")
        axs2.set_ylabel("Liquid Volume, th. sm3")

        axs[0,0].grid()
        axs[0,1].grid()
        axs[1,0].grid()
        axs[1,1].grid()

        plt.close()
        
if __name__ == "__main__":

    window = tk.Tk()

    gui = plot_production(window)

    window.mainloop()
