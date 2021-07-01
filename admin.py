import os

import numpy as np

import openpyxl

from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from tkinter import *
from tkinter import ttk
from tkinter import filedialog

##from ttkwidgets.autocomplete import AutocompleteEntry
##from ttkwidgets.autocomplete import AutocompleteEntryListbox

class schedule():

    def __init__(self,window):

        self.root = window

        self.initUI()

    def initUI(self):

        self.root.title("BHOS-PE Administration")
        
        self.root.configure(background="white")

        menubar = Menu(self.root)
        
        self.root.config(menu=menubar)

        fileMenu = Menu(menubar,tearoff="off")
    
        fileMenu.add_command(label="Open",command=self.open_file)
        fileMenu.add_command(label="Export",command=self.export_schedule)
        fileMenu.add_command(label="Exit",command=self.root.destroy)

        helpMenu = Menu(menubar,tearoff="off")
        helpMenu.add_command(label="Info")
        
        menubar.add_cascade(label="File",menu=fileMenu)
        menubar.add_cascade(label="Help",menu=helpMenu)

        self.set_frame_notebook()
        self.set_frame_courses()

        self.frame_notebook.pack(side=LEFT,expand=1,fill=BOTH)
        self.frame_courses.pack(side=LEFT,expand=1,fill=BOTH)

    def set_frame_notebook(self):

        self.style = ttk.Style(self.root)
        
        configN = {"tabmargins":[2,5,2,0],"tabposition":"wn","background":"white"}

        configT = {"padding":[40, 1, 5, 0],"background":"white"}

        mapT = {"background":[("selected","lightgrey")],
                    "expand":[("selected",[1,1,1,0])]}
        
        settings = {"TNotebook":{"configure":configN},
                "TNotebook.Tab":{"configure":configT,"map":mapT}}
        
        self.style.theme_create("yummy",parent="alt",settings=settings)

        self.style.theme_use("yummy")

        self.frame_notebook = ttk.Notebook(self.root)

        self.frame0 = self.set_frame_notebook_sheet()
        
        self.frame0.pack(fill='both',expand=True)

        self.frame_notebook.add(self.frame0,text="Instructor's Name",compound=RIGHT)

    def set_frame_notebook_sheet(self):

        frame = Frame(self.frame_notebook,width=300,height=200)

        Grid.rowconfigure(frame,1,weight=1)

        Grid.columnconfigure(frame,0,weight=1)
        Grid.columnconfigure(frame,1,weight=1)
        Grid.columnconfigure(frame,2,weight=1)
        Grid.columnconfigure(frame,3,weight=1)
        
        frame.configure(background="white")

        frame.label0 = Label(frame,text="Fall Semester")
        frame.label0.configure(background="white")
        frame.label0.grid(row=0,column=0,columnspan=2,sticky=EW)

        frame.label1 = Label(frame,text="Spring Semester")
        frame.label1.configure(background="white")
        frame.label1.grid(row=0,column=2,columnspan=2,sticky=EW)

        frame.listbox0 = Listbox(frame,width=40,height=20)
        frame.listbox0.grid(row=1,column=0,columnspan=2,sticky=NSEW)
        
        frame.listbox1 = Listbox(frame,width=40,height=20)
        frame.listbox1.grid(row=1,column=2,columnspan=2,sticky=NSEW)

        frame.button_drop0 = Button(frame,text="Drop Selected",
            command=lambda: self.drop_course(frame.listbox0,self.frame_courses.listbox))
        
        frame.button_drop0.configure(background="white")
        frame.button_drop0.grid(row=2,column=0,sticky=EW)
        
        frame.button_clear0 = Button(frame,text="Drop All",
            command=lambda: self.drop_course(frame.listbox0,self.frame_courses.listbox,
                                                    moveall=True))
        
        frame.button_clear0.configure(background="white")
        frame.button_clear0.grid(row=2,column=1,sticky=EW)

        frame.button_drop1 = Button(frame,text="Drop Selected",
            command=lambda: self.drop_course(frame.listbox1,self.frame_courses.listbox))
        
        frame.button_drop1.configure(background="white")
        frame.button_drop1.grid(row=2,column=2,sticky=EW)
        
        frame.button_clear1 = Button(frame,text="Drop All",
            command=lambda: self.drop_course(frame.listbox1,self.frame_courses.listbox,
                                                    moveall=True))
        
        frame.button_clear1.configure(background="white")
        frame.button_clear1.grid(row=2,column=3,sticky=EW)

        frame.button_test = Button(frame,text="Test Instructor's Schedule",
                                   command=self.test_load)
        
        frame.button_test.configure(background="white")
        frame.button_test.grid(row=3,column=0,columnspan=2,sticky=EW)

        frame.button_print = Button(frame,text="Add Instructor's Schedule to Export",
                                   command=self.export_schedule_instructor)
        
        frame.button_print.configure(background="white")
        frame.button_print.grid(row=3,column=2,columnspan=2,sticky=EW)

        return frame

    def set_frame_courses(self):
        
        self.frame_courses = Frame(self.root,width=10,height=20)
        
        Grid.rowconfigure(self.frame_courses,2,weight=1)
        Grid.rowconfigure(self.frame_courses,4,weight=1)

        Grid.columnconfigure(self.frame_courses,0,weight=1)
        Grid.columnconfigure(self.frame_courses,1,weight=1)
        Grid.columnconfigure(self.frame_courses,2,weight=1)
        
        self.frame_courses.configure(background="white")

        self.frame_courses.label = Label(self.frame_courses,text="Courses")
        self.frame_courses.label.configure(background="white")
        self.frame_courses.label.grid(row=0,column=0,columnspan=3,sticky=EW)

##        self.frame_courses.entry = AutocompleteEntryListbox(
##            self.frame_courses,completevalues=[])
##        
##        self.frame_courses.entry.grid(row=1,column=0,columnspan=3,sticky=EW)

        self.frame_courses.listbox = Listbox(self.frame_courses,width=40,height=20)
            
        self.frame_courses.listbox.grid(row=2,column=0,columnspan=3,sticky=NSEW)

        scroll = Scrollbar(self.frame_courses.listbox)

        scroll.config(command=self.frame_courses.listbox.yview)

        self.frame_courses.listbox.config(yscrollcommand=scroll.set)
        
        scroll.pack(side=RIGHT,fill=Y)

        idx = self.frame_notebook.index(self.frame_notebook.select())

        self.frame_courses.button_tofall = \
            Button(self.frame_courses,text="Add to Fall",
                   command=lambda: self.add_course("Fall"))

        self.frame_courses.button_tofall.configure(background="white")
        self.frame_courses.button_tofall.grid(row=3,column=0,sticky=EW)

        self.frame_courses.button_tospring = \
            Button(self.frame_courses,text="Add to Spring",
                   command=lambda: self.add_course("Spring"))

        self.frame_courses.button_tospring.configure(background="white")
        self.frame_courses.button_tospring.grid(row=3,column=1,sticky=EW)

        self.frame_courses.button_extended = Button(self.frame_courses,
                                                    text="Extended View")

        self.frame_courses.button_extended.configure(background="white")
        self.frame_courses.button_extended.grid(row=3,column=2,sticky=EW)

        self.frame_courses.status_text = StringVar()

        self.frame_courses.status = Label(self.frame_courses,
                                          relief="sunken",
                                          textvariable=self.frame_courses.status_text,
                                          anchor=NW)
        
        self.frame_courses.status.configure(background="white")
        self.frame_courses.status.grid(row=4,column=0,columnspan=3,sticky=NSEW)

    def set_input(self):

        wb = openpyxl.load_workbook(self.filepath)

        class teachers: pass
        class courses: pass
        class hours: pass

        sheet_teachers = wb["instructors"]
        sheet_lectures = wb["lectures"]

        teachers.tuple = tuple(sheet_teachers.iter_cols(min_row=1,max_row=18,
                                     min_col=1,max_col=4,values_only=True))

        courses.tuple = tuple(sheet_lectures.iter_cols(min_row=2,max_row=47,
                                     min_col=2,max_col=6,values_only=True))
        
        hours.tuple = tuple(sheet_lectures.iter_rows(min_row=2,max_row=47,
                                     min_col=7,max_col=22,values_only=True))

        self.teachers = teachers
        self.courses = courses
        self.hours = hours

        self.set_teachers()
        self.set_courses()
        self.set_hours()
        
    def set_teachers(self):
        
        for i in range(len(self.teachers.tuple)):
            
            header = self.teachers.tuple[i][0]
            body = np.array(self.teachers.tuple[i][1:])
            
            setattr(self.teachers,header,body)

        name = np.char.add(self.teachers.first," ")

        self.teachers.fullname = np.char.add(name,self.teachers.last)

        for i,name in enumerate(self.teachers.fullname):

            framename = "frame"+str(i)

            frame = self.set_frame_notebook_sheet()

            setattr(self,framename,frame)
            
            getattr(self,framename).pack(fill='both',expand=True)

            self.frame_notebook.add(getattr(self,framename),text=name,compound=RIGHT)

    def set_courses(self):
        
        for i in range(len(self.courses.tuple)):
            
            header = self.courses.tuple[i][0]
            body = np.array(self.courses.tuple[i][1:])
            
            setattr(self.courses,header,body)

        name = np.char.add(self.courses.code,"-")
        
        name = np.char.add(name,self.courses.semester_number.astype(str))

        name = np.char.add(name,"-")
        
        self.courses.description = np.char.add(name,self.courses.name_ENG)

        for entry in self.courses.description:
            self.frame_courses.listbox.insert(END,entry)

##        print(list(self.courses.description))
##
##        self.frame_courses.entry = AutocompleteEntryListbox(
##            self.frame_courses,completevalues=list(self.courses.description))
##
##        self.frame_courses.entry.grid(row=1,column=0,columnspan=3,sticky=EW)

##        for entry in self.courses.description:
##            self.frame_courses.entry.listbox.insert(END,entry)
##        self.frame_courses.entry.set_completion_list(list(self.courses.description))

    def set_hours(self):

        self.hours.definition = self.hours.tuple[0]

        self.hours.nparray = np.array(self.hours.tuple[1:])
        
        self.hours.total = self.hours.nparray.sum(axis=1)

    def open_file(self):

        self.filepath = filedialog.askopenfilename(
            title = "Select a File",
            initialdir = os.getcwd(),
            filetypes = (("Excel files","*.xl*"),("All files","*.*")))

        if not self.filepath:
            return

        for tabid in self.frame_notebook.tabs():
            self.frame_notebook.forget(tabid)
            
        self.frame_courses.listbox.delete(0,END)
        
        self.set_input()

        status = "Opened \""+self.filepath+"\"."
        
        self.frame_courses.status_text.set(status)

    def drop_course(self,frombox,tobox,moveall=False):

        if moveall:
            
            for entry in frombox.get(0,END):
                tobox.insert(END,entry)
                
            frombox.delete(0,END)
        
        elif frombox.curselection():
            
            tobox.insert(END,frombox.get(frombox.curselection()))
            
            frombox.delete(frombox.curselection())

    def add_course(self,semester):

        idx = self.frame_notebook.index(self.frame_notebook.select())

        frameID = "frame"+str(idx)

        frombox = self.frame_courses.listbox

        if semester == "Fall":
            tobox = getattr(self,frameID).listbox0
        elif semester == "Spring":
            tobox = getattr(self,frameID).listbox1

        if frombox.curselection():

            tobox.insert(END,frombox.get(frombox.curselection()))
            
            frombox.delete(frombox.curselection())

    def test_load(self):

        idx = self.frame_notebook.index(self.frame_notebook.select())

        frameID = "frame"+str(idx)

        f0 = getattr(self,frameID).listbox0.get(ACTIVE)
        f1 = getattr(self,frameID).listbox1.get(ACTIVE)

        if not f0 and not f1:
            status = "Total hours is "+str(0)+"."
            self.frame_courses.status_text.set(status)
            return

        s0 = getattr(self,frameID).listbox0.get(0,END)
        s1 = getattr(self,frameID).listbox1.get(0,END)

        annual_load = np.array([np.array(s0+s1)]).T

        idy = np.argwhere(np.array([self.courses.description])==annual_load)[:,1]
            
        hours = self.hours.nparray[idy,:]

        hours_sum = hours.sum(axis=0)

        status = "Total hours is "+str(int(hours_sum[8]))+"."

        self.frame_courses.status_text.set(status)

    def export_schedule_instructor(self):

        idx = self.frame_notebook.index(self.frame_notebook.select())

        frameID = "frame"+str(idx)

        f0 = getattr(self,frameID).listbox0.get(ACTIVE)
        f1 = getattr(self,frameID).listbox1.get(ACTIVE)

        if not f0 and not f1:
            status = "No course is selected."
            self.frame_courses.status_text.set(status)
            return

        s0 = getattr(self,frameID).listbox0.get(0,END)
        s1 = getattr(self,frameID).listbox1.get(0,END)

        annual_load = np.array([np.array(s0+s1)]).T

        idy = np.argwhere(np.array([self.courses.description])==annual_load)[:,1]
            
        hour_rows = self.hours.nparray[idy,:]

        if not hasattr(self,"wb"):
            self.wb = openpyxl.Workbook()

        ws = self.wb.create_sheet(self.teachers.fullname[idx])

        names = list(np.array(["code","name"]))
        hours = list(self.hours.definition)

        ws.append(names+hours)

        names_colnum = len(names)
        hours_colnum = len(hours)

        for i in range(names_colnum+1,names_colnum+hours_colnum+1):

            ws.cell(1,i).alignment = Alignment(textRotation=90)

            ws.column_dimensions[get_column_letter(i)].width = 4

        for i,hour_row in enumerate(hour_rows):

            code = self.courses.code[idy[i]]
            name = self.courses.name_AZE[idy[i]]

            ws.append(list(np.array([code,name]))+list(hour_row))

        status = "Saved " + self.teachers.fullname[idx] + "'s schedule to export"
        
        self.frame_courses.status_text.set(status)

    def export_schedule(self):

        if not hasattr(self,"wb"):
            status = "No schedule is added to export."
            self.frame_courses.status_text.set(status)
            return

        exportfilepath = filedialog.asksaveasfilename(
            title = "Export as",
            initialdir = os.getcwd(),
            defaultextension = ".xlsx",
            filetypes=(("Excel Workbook","*.xlsx"),))
        
        self.wb.save(filename=exportfilepath)

        status = "Saved to \""+exportfilepath+"\"."
        
        self.frame_courses.status_text.set(status)
        
if __name__ == "__main__":
    
    window = Tk()
    
    gui = schedule(window)
                                                                  
##    window.geometry("1100x500")

##    Grid.rowconfigure(window, 0, weight=1)
##    Grid.columnconfigure(window, 0, weight=1)

##    frame = Frame(window)
##
##    Grid.columnconfigure(frame, 0, weight=1)
##    Grid.columnconfigure(frame, 1, weight=1)
##
##    Grid.rowconfigure(frame, 0, weight=1)
##
##    listbox1 = Listbox(frame,width=40,height=20)
##    listbox2 = Listbox(frame,width=80,height=20)
##
##    listbox1.grid(row=0,column=0,sticky=N+E+W+S)
##    listbox2.grid(row=0,column=1,sticky=N+E+W+S)

##    frame.grid(row=0,column=0,sticky=EW)

##    frame.pack(expand=1,fill=BOTH)

    window.mainloop()
