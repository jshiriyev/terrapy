import os

import csv
import openpyxl
import sys

import numpy as np

from PyQt5 import QtCore, QtGui, QtWidgets

inputfile0 = 'C:\\Users\\javid.s\\Desktop\\BHR-076.txt'
inputfile1 = 'C:\\Users\\javid.s\\Desktop\\BHR_VI_SCHEDULE.INC'

outputfile = 'C:\\Users\\javid.s\\Desktop\\BHR_VI_SCHEDULE_V2.INC'

def setup(filename,sheets,*args):

    _,file_extension = os.path.splitext(filename)

    if file_extension == ".csv":
        readcsv(filename,sheets,*args)
    elif file_extension == ".xlsx":
        readxlsx(filename,sheets,*args)

def readcsv(filename,sheets,*args):

    with open(filename) as csvfile:
        reader = list(csv.reader(csvfile))
        reader = np.array(reader)
        num_row, num_col = reader.shape
        if sheets['dataTypes']=='col':
            tuples = tuple(reader.T)
        elif sheets['dataTypes']=='row':
            tuples = tuple(reader)
        setparent(tuples,*args)
        
def readxlsx(filename,sheets,*args):
    all_tuples = ()
    wb = openpyxl.load_workbook(filename)
    for i,sheetname in enumerate(sheets["names"]):
        sheet = wb[sheetname]
        if sheets['dataTypes'][i]=='col':
            tuples = tuple(sheet.iter_cols(max_col=sheets['num_cols'][i],values_only=True))
        elif sheets['dataTypes'][i]=='row':
            tuples = tuple(sheet.iter_rows(max_col=sheets['num_cols'][i],values_only=True))
        all_tuples = all_tuples+tuples
    setparent(all_tuples,*args)

def setparent(tuples,*args):
    for i in range(len(tuples)):
        for arg in args:
            if tuples[i][0] == arg.__name__:
                setchild(tuples[i],arg)

def setchild(atuple,obj):
    array = np.array(atuple[3:]).astype('float64')
    setattr(obj,atuple[1]+'_unit',atuple[2])
    setattr(obj,atuple[1],array)

def skiptoline(file_read,keyword,file_written=None):
    
    while True:
        
        line = next(file_read)

        if file_written is not None:
            file_written.write(line)
        
        if line.split('/')[0].strip() == keyword:
            break

def task0():

    """

    The task was to get dates of well from the first input file
    (inputfile0) and modify the production data of input file 2
    (inputfile1) so that the specified well will have new
    production data.

    """
    
    with open(inputfile0) as readfile:
        
        rdate = np.array([])
        rcopy = np.array([])

        rlines = readfile.readlines()

        for rlinenumber,rline in enumerate(rlines):

            if rlinenumber>0:

                alist = rline.split('\t')

                wellname = alist[0]

                date_list = alist[1].split('-')

                if date_list[1]=='01':
                    date_newformat = date_list[2]+' JAN '+date_list[0]
                elif date_list[1]=='02':
                    date_newformat = date_list[2]+' FEB '+date_list[0]
                elif date_list[1]=='03':
                    date_newformat = date_list[2]+' MAR '+date_list[0]
                elif date_list[1]=='04':
                    date_newformat = date_list[2]+' APR '+date_list[0]
                elif date_list[1]=='05':
                    date_newformat = date_list[2]+' MAY '+date_list[0]
                elif date_list[1]=='06':
                    date_newformat = date_list[2]+' JUN '+date_list[0]
                elif date_list[1]=='07':
                    date_newformat = date_list[2]+' JUL '+date_list[0]
                elif date_list[1]=='08':
                    date_newformat = date_list[2]+' AUG '+date_list[0]
                elif date_list[1]=='09':
                    date_newformat = date_list[2]+' SEP '+date_list[0]
                elif date_list[1]=='10':
                    date_newformat = date_list[2]+' OCT '+date_list[0]
                elif date_list[1]=='11':
                    date_newformat = date_list[2]+' NOV '+date_list[0]
                elif date_list[1]=='12':
                    date_newformat = date_list[2]+' DEC '+date_list[0]

                rdate = np.append(rdate,date_newformat)
                    
                days = alist[2]
                condensate = alist[3]
                gas = alist[4]
                water = alist[5]

                if rlinenumber+1==len(rlines):
                    copied_line = '\t\''+wellname+'\'\tSTOP\tGRAT\t'+condensate+'\t'+water+'\t'+gas+' /\n'
                else:
                    copied_line = '\t\''+wellname+'\'\tOPEN\tGRAT\t'+condensate+'\t'+water+'\t'+gas+' /\n'

                rcopy = np.append(rcopy,copied_line)

            
    with open(inputfile1) as rewrittenfile:

        with open(outputfile,"w") as writtenfile:

            for i,date in enumerate(rdate):

                skiptoline(rewrittenfile,'DATES',writtenfile)
                skiptoline(rewrittenfile,rdate[i],writtenfile)
                skiptoline(rewrittenfile,'WCONHIST',writtenfile)
                
                while True:
                    
                    wline = next(rewrittenfile)
                    
                    if wline.split('/')[0].strip().split(' ')[0] == '\'BHR_76\'':
                        writtenfile.write(rcopy[i])
                        break
                    else:
                        writtenfile.write(wline)
                        
            while True:

                try:
                    
                    copiedline = next(rewrittenfile)
                    writtenfile.write(copiedline)
                    
                except:
                    
                    break

class Window(QtWidgets.QMainWindow):

    #itemEdited = QtCore.pyqtSignal(item,column)
    
    def __init__(self,filepath):
        
        super(Window,self).__init__()

        self.resize(1200,600)

        filepath = "C:\\Users\\Cavid\\Documents\\bhospy\\interfaces\\sample.csv"
        
        self.filepath = filepath
        self.treeView = DataViewer(parent=self)
        self.treeView.loadCSV(self.filepath)

        self.treeView.autoColumnWidth()

        #print(dir(self.treeView))
        
        self.centralwidget = QtWidgets.QWidget(self)
        self.verticalLayout = QtWidgets.QVBoxLayout(self.centralwidget)
        self.verticalLayout.addWidget(self.treeView)
        self.setCentralWidget(self.centralwidget)
        
        self.editedFlag = True

        self.show()   

    def closeEvent(self,event):
        
        if self.treeView.editedFlag:
            choice = QtWidgets.QMessageBox.question(self,'Quit',
                "Do you want to save changes?",QtWidgets.QMessageBox.Yes | 
                QtWidgets.QMessageBox.No)

            if choice == QtWidgets.QMessageBox.Yes:
                self.treeView.writeCSV(self.filepath)

        event.accept()

class DataViewer(QtWidgets.QTreeView):

    def __init__(self,parent=None):
        super(DataViewer,self).__init__(parent)

        #print(dir(self))

        self.setSortingEnabled(True)

        self.model = QtGui.QStandardItemModel()
        self.setModel(self.model)

        #self.setColumnWidth(0,800)

        self.editedFlag = False

        self.model.itemChanged.connect(self.changeFlag)

    def changeFlag(self):
        self.editedFlag = True

    def mouseDoubleClickEvent(self,event):

        index = self.indexAt(event.pos())
        #print(index.row(),index.column())

        #item = self.model.item(index.row(),index.column())
        #print(item.data(QtCore.Qt.DisplayRole))
        #item.setEditable(True)

        if index.row()>-1 and index.column()>-1:
            self.edit(index)

    def loadCSV(self,filepath):

        csvcontent = list(csv.reader(open(filepath)))

        #self.treeView.setHeaderLabels(csvcontent[0])

        self.model.setHorizontalHeaderLabels(csvcontent[0])

        for line in csvcontent[1:]:
            items = [
                QtGui.QStandardItem(text)
                for text in line
                ]
            self.model.appendRow(items)

    def writeCSV(self,filepath):

        with open(filepath,"w") as fileOutput:
            writer = csv.writer(fileOutput)

            headers = [
                self.model.horizontalHeaderItem(columnNumber).text()
                for columnNumber in range(self.model.columnCount())
            ]

            writer.writerow(headers)

            for rowNumber in range(self.model.rowCount()):
                fields = [
                    self.model.data(
                        self.model.index(rowNumber,columnNumber),
                        QtCore.Qt.DisplayRole
                    )
                    for columnNumber in range(self.model.columnCount())
                ]
                writer.writerow(fields)

    def autoColumnWidth(self):
        
        for columnNumber in range(self.model.columnCount()):
            self.resizeColumnToContents(columnNumber)
                
if __name__ == "__main__":

    app = QtWidgets.QApplication(sys.argv)
##    window = Window(args.filepath)
    window = Window(os.curdir)
    sys.exit(app.exec_())

##    class reservoir: pass
##    class well: pass
##    class time: pass
##
##    sheets = {
##        "num_cols": 5,
##        "dataTypes": "col"
##        }
##
##    setup("sample.csv",sheets,reservoir)
##
##    print(reservoir.porosity)
##    
##    sheets = {
##        "names": ["build_up_test","parameters"],
##        "num_cols": [3,4],
##        "dataTypes": ["col","row"]
##        }
##
##    readxlsx('sample.xlsx',sheets,time,well)
