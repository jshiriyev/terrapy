import datetime

import dateutil.parser as parser

import inspect

import os
import re

import tkinter as tk

from tkinter import ttk
from tkinter import filedialog
from tkinter import font as tkfont

from ttkwidgets.autocomplete import AutocompleteEntryListbox

from matplotlib import pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

import numpy as np

import openpyxl

class manager():

    special_extensions = [".db",".xlsx"]

    def __init__(self,filepath=None,skiplines=1,headerline=None,comment="--",endline="/",endfile="END",**kwargs):

        if filepath is None:
            return

        self.filepath = filepath

        self.skiplines = skiplines

        if headerline is None:
            self.headerline = skiplines-1
        elif headerline<skiplines:
            self.headerline = headerline
        else:
            self.headerline = skiplines-1

        self.comment = comment
        self.endline = endline
        self.endfile = endfile

        self.filename = os.path.split(self.filepath)[1]
        self.extension = os.path.splitext(self.filepath)[1]

        if any([self.extension==extension for extension in self.special_extensions]):
            self.read_special(**kwargs)
        else:
            self.read_plain()

        self.title = []

        for _ in range(self.skiplines):
            self.title.append(self._running.pop(0))

        num_cols = len(self._running[0])

        if self.skiplines==0:
            self._headers = ["Column #"+str(index) for index in range(num_cols)]
        elif skiplines!=0:
            self._headers = self.title[self.headerline]

        self.headers = self._headers

        nparray = np.array(self._running).T

        self._running = [np.asarray(column) for column in nparray]

        self.running = [np.asarray(column) for column in self._running]

    def read_plain(self):

        # While looping inside the file it does not read lines:
        # - starting with comment phrase, e.g., comment = "--"
        # - after the end of line phrase, e.g., endline = "/"
        # - after the end of file keyword e.g., endfile = "END"

        self._running = []

        with open(self.filepath,"r") as text:

            for line in text:

                line = line.split('\n')[0].strip()

                if self.endline is not None:
                    line = line.strip(self.endline)

                if line=="":
                    continue

                if self.comment is not None:
                    if line[:len(self.comment)] == self.comment:
                        continue

                if self.endfile is not None:
                    if line[:len(self.endfile)] == self.endfile:
                        break

                self._running.append([line])

    def read_special(self,sheetname=None,min_row=1,min_col=1,max_row=None,max_col=None):

        if self.extension == ".xlsx":

            wb = openpyxl.load_workbook(self.filepath,read_only=True)

            lines = wb[sheetname].iter_rows(min_row=min_row,min_col=min_col,
                max_row=max_row,max_col=max_col,values_only=True)

            self._running = [list(line) for line in lines]

            wb._archive.close()
            return

    def set_subheaders(self,header_index=None,header=None,regex=None,regex_builtin="INC_HEADERS",title="SUB-HEADERS"):

        nparray = np.array(self._running[header_index])

        if regex is None and regex_builtin=="INC_HEADERS":
            regex = r'^[A-Z]+$'                         #for strings with only capital letters no digits
        elif regex is None and regex_builtin=="INC_DATES":
            regex = r'^\d{1,2} [A-Z]{3} \d{2}\d{2}? $'   #for strings with [1 or 2 digits][space][3 capital letters][space][2 or 4 digits], e.g. DATES

        vmatch = np.vectorize(lambda x: bool(re.compile(regex).match(x)))

        match_index = vmatch(nparray)

        firstocc = np.argmax(match_index)

        lower = np.where(match_index)[0]
        upper = np.append(lower[1:],nparray.size)

        repeat_count = upper-lower-1

        match_content = nparray[match_index]

        nparray[firstocc:][~match_index[firstocc:]] = np.repeat(match_content,repeat_count)

        self._headers.insert(header_index,title)
        self._running.insert(header_index,np.asarray(nparray))

        for index,column in enumerate(self._running):
            self._running[index] = np.array(self._running[index][firstocc:][~match_index[firstocc:]])

        self.headers = self._headers
        self.running = [np.asarray(column) for column in self._running]

    def texttocolumn(self,header_index=None,header=None,deliminator=None,maxsplit=None):

        if header_index is None:
            header_index = self._header.index(header)

        header_string = self._headers[header_index]
        # header_string = re.sub(deliminator+'+',deliminator,header_string)

        headers = header_string.split(deliminator)

        if maxsplit is not None:
            for index in range(maxsplit):
                if len(headers)<index+1:
                    headers.append("col ##"+string(header_index+index))

        column_to_split = np.asarray(self._running[header_index])

        running = []

        for index,string in enumerate(column_to_split):

            # string = re.sub(deliminator+'+',deliminator,string)
            row = np.char.split(string,deliminator).tolist()

            if maxsplit is not None:
                while len(row)<maxsplit:
                    row.append("")

            running.append(row)

        running = np.array(running,dtype=str).T

        self._headers.pop(header_index)
        self._running.pop(header_index)

        for header,column in zip(headers,running):
            self._headers.insert(header_index,header)
            self._running.insert(header_index,column)
            header_index += 1

        # line = re.sub(r"[^\w]","",line)
        # line = "_"+line if line[0].isnumeric() else line
        # vmatch = np.vectorize(lambda x:bool(re.compile('[Ab]').match(x)))
        
        self.headers = self._headers
        self.running = [np.asarray(column) for column in self._running]

    def columntotext(self,header_new,header_indices=None,headers=None,string=None):

        if header_indices is None:
            header_indices = [self._headers.index(header) for header in headers]

        if string is None:
            string = ("{} "*len(header_indices)).strip()

        vprint = np.vectorize(lambda *args: string.format(*args))

        column_new = [np.asarray(self._running[index]) for index in header_indices]

        column_new = vprint(*column_new)

        self._headers.append(header_new)
        self._running.append(column_new)

        self.headers = self._headers
        self.running = [np.asarray(column) for column in self._running]

    def astype(self,header_index=None,header=None,dtype=None):

        if header_index is None:
            header_index = self._headers.index(header)

        if inspect.isclass(dtype):
            vdate = np.vectorize(lambda x: dtype(x))
        elif type(dtype)==str:
            if type(self._running[header_index][0])==datetime.datetime:
                vdate = np.vectorize(lambda x: x.strftime(dtype))
            elif any([type(self._running[header_index][0])==class_ for class_ in [str,np.str_,np.str]]):
                vdate = np.vectorize(lambda x: parser.parse(x).strftime(dtype))

        self._running[header_index] = vdate(self._running[header_index])

        self.running[header_index] = np.asarray(self._running[header_index])

    def upper(self,header_index=None,header=None):

        if header_index is None:
            header_index = self._headers.index(header)

        self._running[header_index] = np.char.upper(self._running[header_index])

    def set_column(self,column,header_index=None,header_new=None):
        
        if header_index is None:
            if header_new is None:
                header_new = "Col ##"+str(len(self._headers))
            self._headers.append(header_new)
            self.headers = self._headers
            self._running.append(column)
            self.running.append(np.asarray(self._running[-1]))
        else:
            self._running[header_index] = column
            self.running[header_index] = np.asarray(self._running[header_index])

    def set_rows(self,row,row_indices=None):
        
        if row_indices is None:
            for index,column in enumerate(self._running):
                self._running[index] = np.append(column,row[index])
        else:
            for index, _ in enumerate(self._running):
                self._running[index][row_indices] = row[index]

        self.running = [np.asarray(column) for column in self._running]

    def get_rows(self,row_indices):

        if type(row_indices)==int:
            row_indices = [row_indices]

        rows = [[column[index] for column in self._running] for index in row_indices]
        
        return rows

    def get_columns(self,header_indices=None,headers=None,inplace=False):

        if header_indices is None:
            header_indices = [self._headers.index(header) for header in headers]

        if inplace:
            self._headers = [self._headers[index] for index in header_indices]
            self.headers = self._headers
            self._running = [self._running[index] for index in header_indices]
            self.running = [np.asarray(column) for column in self._running]
        else:
            self.headers = [self._headers[index] for index in header_indices]
            self.running = [np.asarray(self._running[index]) for index in header_indices]

    def del_rows(self,row_indices,inplace=False):

        all_rows = np.array([np.arange(self._running[0].size)])

        row_indices = np.array(row_indices).reshape((-1,1))

        comp_mat = all_rows==row_indices

        keep_index = ~np.any(comp_mat,axis=0)

        if inplace:
            self._running = [column[keep_index] for column in self._running]
            self.running = [np.asarray(column) for column in self._running]
        else:
            self.running = [np.asarray(column[keep_index]) for column in self._running]

    def sort(self,header_indices=None,headers=None,reverse=False,inplace=False,returnFlag=False):

        if header_indices is None:
            header_indices = [self.headers.index(header) for header in headers]

        columns = [self._running[index] for index in header_indices]

        columns.reverse()

        sort_index = np.lexsort(columns)

        if reverse:
            sort_index = np.flip(sort_index)

        if inplace:
            self._running = [column[sort_index] for column in self._running]
            self.running = [np.asarray(column) for column in self._running]
        else:
            self.running = [np.asarray(column[sort_index]) for column in self._running]

        if returnFlag:
            return sort_index

    def filter(self,header_index=None,header=None,keywords=None,regex=None,inplace=False):

        if header_index is None:
            header_index = self._headers.index(header)

        if keywords is not None:
            match_array = np.array(keywords).reshape((-1,1))
            match_index = np.any(self._running[header_index]==match_array,axis=0)
        else:
            match_vectr = np.vectorize(lambda x:bool(re.compile(regex).match(x)))
            match_index = match_vectr(self._running[header_index])

        if inplace:
            self._running = [column[match_index] for column in self._running]
            self.running = [np.asarray(column) for column in self._running]
        else:
            self.running = [np.asarray(column[match_index]) for column in self._running]

    def write_to(self,filepath,header_indices=None,headers=None,string=None,**kwargs):

        output_extension = os.path.splitext(filepath)[1]

        if any([output_extension==extension for extension in self.special_extensions]):
            self.write_special(filepath,**kwargs)
            return

        if header_indices is None:
            header_indices = [self._headers.index(header) for header in headers]

        if string is None:
            string = ("{}\t"*len(header_indices))[:-1]+"\n"

        vprint = np.vectorize(lambda *args: string.format(*args))

        columns = [np.asarray(self._running[index]) for index in header_indices]

        with open(filepath,"w",encoding='utf-8') as wfile:
            for line in vprint(*columns):
                wfile.write(line)

    def write_special(self,filepath,**kwargs):

        if extension == ".xlsx":
            wb = openpyxl.Workbook()
            sheet = wb.active
            if sheet_title is not None:
                sheet.title = sheet_title
            for line in running:
                sheet.append(line)
            wb.save(filepath)
            return

class tree():

    def __init__(self,dirpath):

        self.dirpath = dirpath

    def draw(self,window,func=None):

        self.root = window

        self.scrollbar = ttk.Scrollbar(self.root)

        self.tree = ttk.Treeview(self.root,show="headings tree",selectmode="browse",yscrollcommand=self.scrollbar.set)

        self.tree.pack(side=tk.LEFT,expand=1,fill=tk.BOTH)

        self.scrollbar.pack(side=tk.LEFT,fill=tk.Y)

        self.scrollbar.config(command=self.tree.yview)

        self.tree.bind("<Button-1>",lambda event: self.set_path(func,event))

        self.refill()

    def refill(self):

        self.tree.heading("#0",text="")

        self.tree.delete(*self.tree.get_children())

        iterator = os.walk(self.dirpath)

        parents_name = []
        parents_link = []

        counter  = 0

        while True:

            try:
                root,dirs,files = next(iterator)
            except StopIteration:
                break

            if counter==0:
                dirname = os.path.split(root)[1]
                self.tree.heading("#0",text=dirname,anchor=tk.W)
                parents_name.append(root)
                parents_link.append("")
            
            parent = parents_link[parents_name.index(root)]

            for directory in dirs:
                link = self.tree.insert(parent,'end',iid=counter,text=directory)
                counter += 1

                parents_name.append(os.path.join(root,directory))
                parents_link.append(link)

            for file in files:            
                self.tree.insert(parent,'end',iid=counter,text=file)
                counter += 1

    def set_path(self,func=None,event=None):

        if event is not None:
            region = self.tree.identify("region",event.x,event.y)
        else:
            return

        if region!="tree":
            return

        item = self.tree.identify("row",event.x,event.y)

        path = self.tree.item(item)['text']

        while True:

            item = self.tree.parent(item)

            if item:
                path = os.path.join(self.tree.item(item)['text'],path)
            else:
                path = os.path.join(self.dirpath,path)
                break

        if func is not None:
            func(path)
        else:
            print(path)

class table(manager):

    def __init__(self,filepath=None,headers=None,**kwargs):

        super().__init__(filepath,**kwargs)

        if filepath is None:

            self._headers = headers
            self._running = [np.array([]) for _ in self._headers]

            self.headers = self._headers
            self.running = [np.asarray(column) for column in self._running]

    def draw(self,window,func=None):

        self.root = window

        self.scrollbar = tk.Scrollbar(self.root)

        self.columns = ["#"+str(idx) for idx,_ in enumerate(self.headers,start=1)]

        self.tree = ttk.Treeview(self.root,columns=self.columns,show="headings",selectmode="browse",yscrollcommand=self.scrollbar.set)

        self.sortReverseFlag = [False for column in self.columns]

        for idx,(column,header) in enumerate(zip(self.columns,self.headers)):
            self.tree.column(column,anchor=tk.W,stretch=tk.NO)
            self.tree.heading(column,text=header,anchor=tk.W)

        self.tree.column(self.columns[-1],stretch=tk.YES)

        self.tree.pack(side=tk.LEFT,expand=1,fill=tk.BOTH)

        self.scrollbar.pack(side=tk.LEFT,fill=tk.Y)

        self.scrollbar.config(command=self.tree.yview)

        # self.frame = tk.Frame(self.root,width=50)
        # self.frame.configure(background="white")
        # self.frame.pack(side=tk.LEFT,fill=tk.Y)

        self.tree.bind("<KeyPress-i>",self.addItem)

        # self.button_Add = tk.Button(self.frame,text="Add Item",width=50,command=self.addItem)
        # self.button_Add.pack(side=tk.TOP,ipadx=5,padx=10,pady=(5,1))

        self.tree.bind("<KeyPress-e>",self.editItem)
        self.tree.bind("<Double-1>",self.editItem)

        self.tree.bind("<Delete>",self.deleteItem)

        self.tree.bind("<KeyPress-j>",self.moveDown)
        self.tree.bind("<KeyPress-k>",self.moveUp)

        self.tree.bind("<Button-1>",self.sort_column)

        self.tree.bind("<Control-KeyPress-s>",lambda event: self.saveChanges(func,event))

        # self.button_Save = tk.Button(self.frame,text="Save Changes",width=50,command=lambda: self.saveChanges(func))
        # self.button_Save.pack(side=tk.TOP,ipadx=5,padx=10,pady=(10,1))

        self.root.protocol('WM_DELETE_WINDOW',lambda: self.close_no_save(func))

        self.counter = self.running[0].size

        self.iids = np.arange(self.counter)

        self.added = []
        self.edited = []
        self.deleted = []

        self.refill()

    def refill(self):

        self.tree.delete(*self.tree.get_children())

        rows = np.array(self.running).T.tolist()

        for iid,row in zip(self.iids,rows):
            self.tree.insert(parent="",index="end",iid=iid,values=row)

    def addItem(self,event):

        if hasattr(self,"topAddItem"):
            if self.topAddItem.winfo_exists():
                return

        self.topAddItem = tk.Toplevel()

        self.topAddItem.resizable(0,0)

        for index,header in enumerate(self.headers):
            label = "label_"+str(index)
            entry = "entry_"+str(index)
            pady = (30,5) if index==0 else (5,5)
            setattr(self.topAddItem,label,tk.Label(self.topAddItem,text=header,font="Helvetica 11",width=20,anchor=tk.E))
            setattr(self.topAddItem,entry,tk.Entry(self.topAddItem,width=30,font="Helvetica 11"))
            getattr(self.topAddItem,label).grid(row=index,column=0,ipady=5,padx=(10,5),pady=pady)
            getattr(self.topAddItem,entry).grid(row=index,column=1,ipady=5,padx=(5,10),pady=pady)

        self.topAddItem.entry_0.focus()

        self.topAddItem.button = tk.Button(self.topAddItem,text="Add Item",command=self.addItemEnterClicked)
        self.topAddItem.button.grid(row=index+1,column=0,columnspan=2,ipady=5,padx=15,pady=(15,30),sticky=tk.EW)

        self.topAddItem.button.bind('<Return>',self.addItemEnterClicked)

        self.topAddItem.mainloop()

    def addItemEnterClicked(self,event=None):

        if event is not None and event.widget!=self.topAddItem.button:
            return

        values = []

        for idx,header in enumerate(self.headers):
            entry = "entry_"+str(idx)
            value = getattr(self.topAddItem,entry).get()
            values.append(value)

        self.added.append(self.counter)

        self.set_rows(values)

        self.iids = np.append(self.iids,self.counter)

        self.tree.insert(parent="",index="end",iid=self.counter,values=values)

        self.counter += 1

        self.topAddItem.destroy()

    def editItem(self,event):

        region = self.tree.identify('region',event.x,event.y)

        if region=="separator":
            self.autowidth(event)
            return

        if not(region=="cell" or event.char=="e"):
            return

        if hasattr(self,"topEditItem"):
            if self.topEditItem.winfo_exists():
                return

        if not self.tree.selection():
            return
        else:
            item = self.tree.selection()[0]

        values = self.tree.item(item)['values']

        self.topEditItem = tk.Toplevel()

        self.topEditItem.resizable(0,0)

        for idx,(header,explicit) in enumerate(zip(self.headers,self.headers)):
            label = "label_"+str(idx)
            entry = "entry_"+str(idx)
            pady = (30,5) if idx==0 else (5,5)
            setattr(self.topEditItem,label,tk.Label(self.topEditItem,text=explicit,font="Helvetica 11",width=20,anchor=tk.E))
            setattr(self.topEditItem,entry,tk.Entry(self.topEditItem,width=30,font="Helvetica 11"))
            getattr(self.topEditItem,label).grid(row=idx,column=0,ipady=5,padx=(10,5),pady=pady)
            getattr(self.topEditItem,entry).grid(row=idx,column=1,ipady=5,padx=(5,10),pady=pady)
            getattr(self.topEditItem,entry).insert(0,values[idx])

        self.topEditItem.entry_0.focus()

        self.topEditItem.button = tk.Button(self.topEditItem,text="Save Item Edit",command=lambda: self.editItemEnterClicked(item))
        self.topEditItem.button.grid(row=idx+1,column=0,columnspan=2,ipady=5,padx=15,pady=(15,30),sticky=tk.EW)

        self.topEditItem.button.bind('<Return>',lambda event: self.editItemEnterClicked(item,event))

        self.topEditItem.mainloop()

    def editItemEnterClicked(self,item,event=None):

        if event is not None and event.widget!=self.topEditItem.button:
            return

        values = []

        for idx,header in enumerate(self.headers):
            entry = "entry_"+str(idx)
            value = getattr(self.topEditItem,entry).get()
            values.append(value)

        self.edited.append([int(item),self.tree.item(item)["values"]])

        self.set_rows(values,np.argmax(self.iids==int(item)))

        self.tree.item(item,values=values)

        self.topEditItem.destroy()

    def deleteItem(self,event):

        if not self.tree.selection():
            return
        else:
            item = self.tree.selection()[0]

        self.deleted.append([int(item),self.tree.item(item)["values"]])

        self.del_rows(np.argmax(self.iids==int(item)),inplace=True)

        self.iids = np.delete(self.iids,np.argmax(self.iids==int(item)))

        self.tree.delete(item)

    def autowidth(self,event):

        column = self.tree.identify('column',event.x,event.y)

        index = self.columns.index(column)

        if index==len(self.columns)-1:
            return

        header_char_count = len(self.headers[index])

        vcharcount = np.vectorize(lambda x: len(x))

        if self.running[index].size != 0:
            column_char_count = vcharcount(self.running[index].astype(str)).max()
        else:
            column_char_count = 0

        char_count = max(header_char_count,column_char_count)

        width = tkfont.Font(family="Consolas", size=12).measure("A"*char_count)

        column_width_old = self.tree.column(column,"width")

        self.tree.column(column,width=width)

        column_width_new = self.tree.column(column,"width")

        column_width_last_old = self.tree.column(self.columns[-1],"width")

        column_width_last_new = column_width_last_old+column_width_old-column_width_new

        self.tree.column(self.columns[-1],width=column_width_last_new)

    def moveUp(self,event):
 
        if not self.tree.selection():
            return
        else:
            item = self.tree.selection()[0]

        self.tree.move(item,self.tree.parent(item),self.tree.index(item)-1)

    def moveDown(self,event):

        if not self.tree.selection():
            return
        else:
            item = self.tree.selection()[0]

        self.tree.move(item,self.tree.parent(item),self.tree.index(item)+1)

    def sort_column(self,event):

        region = self.tree.identify('region',event.x,event.y)

        if region!="heading":
            return

        column = self.tree.identify('column',event.x,event.y)

        header_index = self.columns.index(column)

        reverseFlag = self.sortReverseFlag[header_index]

        N = self.running[0].size

        argsort = np.argsort(self.running[header_index])

        if reverseFlag:
            argsort = np.flip(argsort)

        for index,column in enumerate(self.running):
            self.running[index] = column[argsort]

        self.iids = self.iids[argsort]
        # indices = np.arange(N)

        # sort_indices = indices[np.argsort(argsort)]

        self.refill()

        # for item,sort_index in zip(self.iids,sort_indices):
        #     self.tree.move(item,self.tree.parent(item),sort_index)

        self.sortReverseFlag[header_index] = not reverseFlag

    def saveChanges(self,func=None,event=None):

        self.added = []
        self.edited = []
        self.deleted = []

        if func is not None:
            func()

    def close_no_save(self,func=None):

        try:
            for deleted in self.deleted:
                self.set_rows(deleted[1])
        except:
            print("Could not bring back deleted rows ...")

        try:
            for edited in self.edited:
                self.set_rows(edited[1],np.argmax(self.iids==edited[0]))
        except:
            print("Could not bring back editions ...")

        added = [np.argmax(self.iids==add) for add in self.added]

        try:
            self.del_rows(added,inplace=True)
        except:
            print("Could not remove additions ...")

        try:
            if func is not None:
                func()
        except:
            print("Could not run the called function ...")

        self.root.destroy()

class graph(manager):

    def __init__(self,filepath=None,**kwargs):

        super().__init__(filepath,**kwargs)

    def draw(self,window):

        self.root = window

        # configuration of window pane
        self.pane_NS = ttk.PanedWindow(self.root,orient=tk.VERTICAL,width=1000)

        self.body = ttk.Frame(self.root,height=450)
        self.foot = tk.Listbox(self.root,height=5)

        self.pane_NS.add(self.body,weight=1)
        self.pane_NS.add(self.foot,weight=0)

        self.pane_NS.pack(expand=1,fill=tk.BOTH)

        # configuration of top pane
        self.pane_EW = ttk.PanedWindow(self.body,orient=tk.HORIZONTAL)

        self.side = ttk.Frame(self.body)

        self.figure = plt.Figure()
        self.canvas = FigureCanvasTkAgg(self.figure,self.body)

        self.plot = self.canvas.get_tk_widget()

        self.pane_EW.add(self.side,weight=0)
        self.pane_EW.add(self.plot,weight=1)

        self.pane_EW.pack(expand=1,fill=tk.BOTH)

        # configuration of top left pane
        self.pane_ns = ttk.PanedWindow(self.side,orient=tk.VERTICAL,width=300)

        self.searchbox = AutocompleteEntryListbox(self.side,height=250,padding=0)
        self.searchbox.config(completevalues=[],allow_other_values=False)

        self.template = ttk.Frame(self.side,height=200)

        self.template.rowconfigure(0,weight=0)
        self.template.rowconfigure(1,weight=1)

        self.template.columnconfigure(0,weight=1)
        self.template.columnconfigure(1,weight=0)
        self.template.columnconfigure(2,weight=0)

        self.template_label = ttk.Label(self.template,text="Templates")
        self.template_label.grid(row=0,column=0,sticky=tk.EW)

        icon1 = tk.PhotoImage(file=".\\graphics\\Add\\Add-9.png")
        icon2 = tk.PhotoImage(file=".\\graphics\\Edit\\Edit-9.png")
        icon3 = tk.PhotoImage(file=".\\graphics\\Delete\\Delete-9.png")

        self.button1 = ttk.Button(self.template,image=icon1,width=3,command=self.add_template)
        self.button1.grid(row=0,column=1)

        self.button2 = ttk.Button(self.template,image=icon2,width=3,command=self.edit_template)
        self.button2.grid(row=0,column=2)

        self.button3 = ttk.Button(self.template,image=icon3,width=3,command=self.remove_template)
        self.button3.grid(row=0,column=3)

        self.template_listbox = tk.Listbox(self.template,exportselection=False)
        self.template_listbox.grid(row=1,column=0,columnspan=4,sticky=tk.NSEW)

        self.pane_ns.add(self.searchbox,weight=1)
        self.pane_ns.add(self.template,weight=1)

        self.pane_ns.pack(expand=1,fill=tk.BOTH)

    def add_template(self):
        pass
    def remove_template(self):
        pass
    def edit_template(self):
        pass

    def set_template(self,event):

        if not self.listbox_template.curselection(): return
        
        if hasattr(self,"axes"):
            for axis in self.axes:
                self.figure.delaxes(axis)

        self.axes = []

        self.axes.append(self.figure.add_subplot(221))
        self.axes.append(self.axes[0].twinx())
        self.axes.append(self.figure.add_subplot(222))
        self.axes.append(self.axes[2].twinx())
        self.axes.append(self.figure.add_subplot(223))
        self.axes.append(self.axes[4].twinx())
        self.axes.append(self.figure.add_subplot(224))

        self.axes[0].set_xlabel("Date")
        self.axes[2].set_xlabel("Date")
        self.axes[4].set_xlabel("Date")
        self.axes[6].set_xlabel("Date")

        self.axes[0].set_ylabel("Liquid Rate, sm3/day")
        self.axes[2].set_ylabel("Gas Rate, th. sm3/day")
        self.axes[4].set_ylabel("Liquid Rate, sm3/day")
        self.axes[6].set_ylabel("Pressure, Bars")

        self.axes[1].set_ylabel("Liquid Volume, th. sm3")
        self.axes[3].set_ylabel("Surface Gas Volume, mln. sm3")
        self.axes[5].set_ylabel("Liquid Volume, th. sm3")

        self.axes[0].grid()
        self.axes[2].grid()
        self.axes[4].grid()
        self.axes[6].grid()

        self.figure.set_tight_layout(True)

        self.plot.draw()

        status = "Production history match template has been selected."
        self.status.insert(tk.END,status)
        self.status.see(tk.END)
        
    def set_lines(self,event):

        if not self.listbox.curselection(): return

        if not hasattr(self,"axes"):
            status = "No template has been selected."    
            self.status.insert(tk.END,status)
            self.status.see(tk.END)
            return

        if hasattr(self,"lines"):
            for line in self.lines:
                line.remove()
                
        self.lines = []
        
        try: self.lines.append(self.axes[0].plot(data.date,data.Oil_Rate,c='k')[0])
        except: pass
        try: self.lines.append(self.axes[0].plot(data.date,data.Oil_Rate_H,'--',c='g')[0])
        except: pass
        try: self.lines.append(self.axes[1].plot(data.date,data.Oil_Total/1000,c='k')[0])
        except: pass
        try: self.lines.append(self.axes[1].plot(data.date,data.Oil_Total_H/1000,'--',c='g')[0])
        except: pass
        
        try: self.lines.append(self.axes[2].plot(data.date,data.Gas_Rate/1000,c='k')[0])
        except: pass
        try: self.lines.append(self.axes[2].plot(data.date,data.Gas_Rate_H/1000,'--',c='r')[0])
        except: pass
        try: self.lines.append(self.axes[3].plot(data.date,data.Gas_Total/1000000,c='k')[0])
        except: pass
        try: self.lines.append(self.axes[3].plot(data.date,data.Gas_Total_H/1000000,'--',c='r')[0])
        except: pass
        
        try: self.lines.append(self.axes[4].plot(data.date,data.Water_Rate,c='k')[0])
        except: pass
        try: self.lines.append(self.axes[4].plot(data.date,data.Water_Rate_H,'--',c='b')[0])
        except: pass
        try: self.lines.append(self.axes[5].plot(data.date,data.Water_Total/1000,c='k')[0])
        except: pass
        try: self.lines.append(self.axes[5].plot(data.date,data.Water_Total_H/1000,'--',c='b')[0])
        except: pass
        
        try: self.lines.append(self.axes[6].plot(data.date,data.Bottom_Hole_Pressure,c='k')[0])
        except: self.lines.append(self.axes[6].plot(data.date,data.Avg_Pressure,c='k')[0])
        try: self.lines.append(self.axes[6].plot(data.date,data.Bottom_Hole_Pressure_H,'--',c='m')[0])
        except: pass

        for axis in self.axes:
            axis.relim()
            axis.autoscale_view()

        self.figure.set_tight_layout(True)

        self.plot.draw()

if __name__ == "__main__":

    import time

    
    window = tk.Tk()

    gui = graph()

    gui.draw(window)

    # gui = table("instructors.csv")
    # gui.texttocolumn(0,deliminator=",")

    # gui = tree("C:\\Users\\Cavid\\Documents")

    # t0 = time.time()
    # gui.draw(window)
    # t1 = time.time()

    # total = t1-t0
    

    # print(total)

    # gui = table(headers=["Full Name","Position","Contact"])

    # printer = lambda: [print(name) for name in gui.running[0]]

    # gui.draw(window,printer)

    window.mainloop()
