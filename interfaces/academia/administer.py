import os

import numpy as np

import openpyxl

from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

import tkinter as tk

from tkinter import ttk
from tkinter import filedialog

from ttkwidgets.autocomplete import AutocompleteEntryListbox

if __name__ == "__main__":
    import setup

from interfaces.table import table

class schedule():

    def __init__(self,window):

        self.root = window
        self.root.title("BHOS-PE Administration")
        self.root.configure(background="white")

        menubar = tk.Menu(self.root)

        self.root.config(menu=menubar)

        fileMenu = tk.Menu(menubar,tearoff="Off")
        fileMenu.add_command(label="Open",command=self.open)
        fileMenu.add_command(label="Save",command=self.save)
        fileMenu.add_command(label="Save As ...",command=self.saveAs)
        fileMenu.add_separator()
        fileMenu.add_command(label="Import ...",command=self.import_objects)
        fileMenu.add_command(label="Export ...",command=self.export_schedule)
        fileMenu.add_separator()
        fileMenu.add_command(label="Exit",command=self.root.destroy)

        editMenu = tk.Menu(menubar,tearoff="Off")
        editMenu.add_command(label="Instructors",command=self.edit_instructors)
        editMenu.add_command(label="Courses",command=self.edit_courses)

        helpMenu = tk.Menu(menubar,tearoff="Off")
        helpMenu.add_command(label="Info")
        
        menubar.add_cascade(label="File",menu=fileMenu)
        menubar.add_cascade(label="Edit",menu=editMenu)
        menubar.add_cascade(label="Help",menu=helpMenu)

        self.instructors = table(headers=["Full Name","Position"])
        self.courses = table(headers=["Code","Title","Type","Credits","Enrolment"])
        self.classroom = table(headers=["Type","Capacity"])
        self.connectivity = table(headers=["Instructor Name","Course ID","Semester","Classroom"])

        self.frame_notebook = ttk.Notebook(self.root)
        self.frame0 = self.set_frame_notebook_sheet()
        self.frame0.pack(fill='both',expand=True)
        self.frame_notebook.add(self.frame0,text="Instructor's Name",compound=tk.RIGHT)
        self.frame_notebook.pack(side=tk.LEFT,expand=1,fill=tk.BOTH)
        
        self.set_frame_courses()

        self.frame_courses.pack(side=tk.LEFT,expand=1,fill=tk.BOTH)

        """STYLING"""
        self.style = ttk.Style(self.root)

        # self.style.theme_create("yummy",parent="alt",settings=settings)

        # aqua,step,clam,alt,default,classic

        self.style.theme_use("clam")

        self.style.configure("TNotebook",
                             tabposition="wn",
                             background="#FFD580",
                             borderwidt=0) # tabmargins=[2,5,2,0],

        self.style.configure("TNotebook.Tab",
                             background="#FFD580",
                             font=("Helvetica",11),
                             borderwidth=2,
                             relief="flat",
                             foreground="black",
                             width=20,
                             anchor=tk.E) # padding=[40, 1, 5, 0],

        self.style.configure("AutocompleteEntryListbox",background="white")

        self.style.map("TNotebook.Tab",background=[("selected","#CC5500")],foreground=[("selected","white")],
            focuscolor=self.style.configure(".")["background"],expand=[("selected",[0,0,1,1])])
        #,expand=[("selected",[0,0,0,0])],width=[("selected",22)],
        
        """END-OF-STYLING"""  

    def set_frame_notebook_sheet(self):

        frame = tk.Frame(self.frame_notebook,width=300,height=200)

        tk.Grid.rowconfigure(frame,1,weight=1)

        tk.Grid.columnconfigure(frame,0,weight=1)
        tk.Grid.columnconfigure(frame,1,weight=1)
        tk.Grid.columnconfigure(frame,2,weight=1)
        tk.Grid.columnconfigure(frame,3,weight=1)
        
        frame.configure(background="white")

        frame.label0 = ttk.Label(frame,text="Fall Term")
        frame.label0.configure(background="#CC5500",font=("Helvetica", 11),foreground="white",anchor=tk.CENTER,borderwidth=5,relief="groove")
        frame.label0.grid(row=0,column=0,columnspan=2,sticky=tk.EW)

        frame.label1 = ttk.Label(frame,text="Spring Term")
        frame.label1.configure(background="#CC5500",font=("Helvetica", 11),foreground="white",anchor=tk.CENTER,borderwidth=5,relief="groove")
        frame.label1.grid(row=0,column=2,columnspan=2,sticky=tk.EW)

        frame.listbox0 = tk.Listbox(frame,width=40,height=20)
        frame.listbox0.grid(row=1,column=0,columnspan=2,sticky=tk.NSEW)
        
        frame.listbox1 = tk.Listbox(frame,width=40,height=20)
        frame.listbox1.grid(row=1,column=2,columnspan=2,sticky=tk.NSEW)

        frame.button_drop0 = ttk.Button(frame,text="Drop Selected",
            command=lambda: self.drop_course("Fall Term"))
        # frame.button_drop0.configure(background="white",relief="ridge",activebackground='#8b0000')
        frame.button_drop0.grid(row=2,column=0,sticky=tk.EW)
        
        frame.button_clear0 = ttk.Button(frame,text="Drop All",
            command=lambda: self.drop_course("Fall Term",moveall=True))
        # frame.button_clear0.configure(background="white",relief="ridge",activebackground='#345')
        frame.button_clear0.grid(row=2,column=1,sticky=tk.EW)

        frame.button_drop1 = ttk.Button(frame,text="Drop Selected",
            command=lambda: self.drop_course("Spring Term"))
        # frame.button_drop1.configure(background="white",relief="ridge",activebackground='#345')
        frame.button_drop1.grid(row=2,column=2,sticky=tk.EW)
        
        frame.button_clear1 = ttk.Button(frame,text="Drop All",
            command=lambda: self.drop_course("Spring Term",moveall=True))
        # frame.button_clear1.configure(background="white",relief="ridge",activebackground='#345')
        frame.button_clear1.grid(row=2,column=3,sticky=tk.EW)

        frame.button_test = ttk.Button(frame,text="Test Instructor's schedule",command=self.test_load)
        # frame.button_test.configure(background="white",relief="ridge",activebackground='#345')
        frame.button_test.grid(row=3,column=0,columnspan=2,sticky=tk.EW)

        frame.button_print = ttk.Button(frame,text="Add Instructor's schedule to Export",command=self.add_export)
        # frame.button_print.configure(background="white",relief="ridge",activebackground='#345')
        frame.button_print.grid(row=3,column=2,columnspan=2,sticky=tk.EW)

        return frame

    def set_frame_courses(self):
        
        self.frame_courses = tk.Frame(self.root)
        
        tk.Grid.rowconfigure(self.frame_courses,1,weight=1)
        tk.Grid.rowconfigure(self.frame_courses,4,weight=1)

        tk.Grid.columnconfigure(self.frame_courses,0,weight=1)
        tk.Grid.columnconfigure(self.frame_courses,1,weight=1)
        tk.Grid.columnconfigure(self.frame_courses,2,weight=1)

        self.frame_courses.searchbox = AutocompleteEntryListbox(
            self.frame_courses,width=40,height=40,completevalues=[],allow_other_values=False)
        
        self.frame_courses.searchbox.grid(row=1,column=0,columnspan=3,sticky=tk.NSEW)

        ## self.frame_courses.listbox = Listbox(self.frame_courses,width=40,height=20)
        ## self.frame_courses.listbox.grid(row=2,column=0,columnspan=3,sticky=NSEW)

        idx = self.frame_notebook.index(self.frame_notebook.select())

        self.frame_courses.button_tofall = ttk.Button(
            self.frame_courses,text="Add to Fall",command=lambda: self.add_course("Fall Term"))
        # self.frame_courses.button_tofall.configure(background="white",relief="ridge",activebackground='#345')
        self.frame_courses.button_tofall.grid(row=3,column=0,sticky=tk.EW)

        self.frame_courses.button_tospring = ttk.Button(
            self.frame_courses,text="Add to Spring",command=lambda: self.add_course("Spring Term"))
        # self.frame_courses.button_tospring.configure(background="white",relief="ridge",activebackground='#345')
        self.frame_courses.button_tospring.grid(row=3,column=1,sticky=tk.EW)

        self.frame_courses.button_extended = ttk.Button(self.frame_courses,text="Extended View",command=self.extended_view)
        # self.frame_courses.button_extended.configure(background="white",relief="ridge",activebackground='#345')
        self.frame_courses.button_extended.grid(row=3,column=2,sticky=tk.EW)

        self.frame_courses.status = tk.Text(self.frame_courses,width=40,height=10,wrap=tk.CHAR)
        self.frame_courses.status.grid(row=4,column=0,columnspan=3,sticky=tk.NSEW)

        ## self.frame_courses.status_text = StringVar()
        ## self.frame_courses.status = Label(
        ##      self.frame_courses,relief="sunken",anchor=NW,
        ##      textvariable=self.frame_courses.status_text,wraplength=200)
        ## scroll = Scrollbar(self.frame_courses.status)
        ## scroll.configure(command=self.frame_courses.status.yview)
        ## self.frame_courses.status.configure(yscrollcommand=scroll.set) 
        ## scroll.pack(side=RIGHT,fill=Y)

    def import_objects(self):

        self.topImport = tk.Toplevel()

        self.topImport.resizable(0,0)

        control_variable = tk.StringVar(self.topImport)
        control_variable.set("Select an Object")

        self.topImportOptions = ("Instructors", "Courses")

        self.topImportOptionMenu = ttk.Combobox(self.topImport,state="readonly",textvariable=control_variable)
        self.topImportOptionMenu['values'] = self.topImportOptions
        self.topImportOptionMenu['width'] = 40 
        
        self.topImportOptionMenu.pack()

        self.topImportButton = ttk.Button(self.topImport,text="Submit",command=lambda: self.set_import_path(control_variable.get()))
        self.topImportButton.config(width=40)

        self.topImportButton.pack()

        self.topImport.mainloop()

    def set_import_path(self,object_name):

        self.topImport.destroy()

        if not any([object_name==option for option in self.topImportOptions]): return

        filepath = filedialog.askopenfilename(
            title = "Select a File",
            initialdir = os.getcwd(),
            filetypes = (("Databases","*.db"),
                         ("CSV Files","*.csv"),
                         ("Excel Files","*.xl*"),
                         ("All Files","*")))

        if not filepath: return

        if object_name=="Instructors":
            self.instructors = table(filepath,skiplines=1)
            self.instructors.full_name = self.instructors.get_concatenated("first_name","last_name",deliminator=" ")
            self.set_notebook()
        elif object_name=="Courses":
            self.courses = table(filepath,sheetname="courses",skiplines=2,max_col=4)
            self.set_courses()

        status = "Imported \""+filepath+"\".\n"
        
        self.frame_courses.status.insert(tk.END,status)
        self.frame_courses.status.see(tk.END)
        
    def set_notebook(self):

        for tabid in self.frame_notebook.tabs():
            self.frame_notebook.forget(tabid)

        self.instructors.full_name = self.instructors.get_concatenated("first_name","last_name",deliminator=" ")

        for idx,name in enumerate(self.instructors.full_name):

            framename = "frame"+str(idx)

            frame = self.set_frame_notebook_sheet()

            setattr(self,framename,frame)
            
            getattr(self,framename).pack(fill='both',expand=True)

            self.frame_notebook.add(getattr(self,framename),text=name,compound=tk.RIGHT)

    def set_courses(self):

        self.frame_courses.searchbox.content = []

        self.frame_courses.searchbox.configure(
            completevalues=self.frame_courses.searchbox.content,allow_other_values=False)

        self.courses.description = self.courses.get_concatenated("code","name_eng",deliminator=" ")

        self.frame_courses.searchbox.content = self.courses.description

        self.frame_courses.searchbox.content.sort()

        self.frame_courses.searchbox.configure(
            completevalues=self.frame_courses.searchbox.content,allow_other_values=True)

    def drop_course(self,semester,moveall=False):

        idx = self.frame_notebook.index(self.frame_notebook.select())

        instructor = self.instructors.full_name[idx]

        frameID = "frame"+str(idx)

        if semester == "Fall Term":
            frombox = getattr(self,frameID).listbox0
        elif semester == "Spring Term":
            frombox = getattr(self,frameID).listbox1

        tobox = self.frame_courses.searchbox

        if not frombox.get(0,tk.END):
            status = "No course to drop.\n"
            self.frame_courses.status.insert(tk.END,status)
            self.frame_courses.status.yview(tk.END)
            return
        elif moveall:
            dropped = list(frombox.get(0,tk.END))
            status = "Dropped all courses.\n"
            frombox.delete(0,tk.END)
        elif frombox.curselection():
            dropped = [frombox.get(frombox.curselection())]
            status = "Dropped selected course.\n"
            frombox.delete(frombox.curselection())
        else:
            status = "No course is selected.\n"
            self.frame_courses.status.insert(tk.END,status)
            self.frame_courses.status.yview(tk.END)
            return

        body_rows = np.array(self.connectivity.body_cols).T.tolist()
        
        for course in dropped:
            body_rows.remove([instructor,course,semester])
            self.connectivity.num_rows -= 1
            tobox.content.append(course)

        tobox.content.sort()
        tobox.configure(completevalues=tobox.content)

        self.connectivity.body_cols = np.array(body_rows).T.tolist()
            
        self.frame_courses.status.insert(tk.END,status)
        self.frame_courses.status.yview(tk.END)
            
    def add_course(self,semester):

        idx = self.frame_notebook.index(self.frame_notebook.select())

        instructor = self.instructors.full_name[idx]

        frameID = "frame"+str(idx)

        frombox = self.frame_courses.searchbox

        if semester == "Fall Term":
            tobox = getattr(self,frameID).listbox0
        elif semester == "Spring Term":
            tobox = getattr(self,frameID).listbox1

        if frombox.listbox.curselection(): 
            course = frombox.listbox.get(frombox.listbox.curselection())
            frombox.content.remove(course)
            frombox.configure(completevalues=frombox.content)
            frombox.entry.delete(0,tk.END)
            status = "The course is added.\n"
        else:
            status = "Select a course to add.\n"
            return

        body_rows = np.array(self.connectivity.body_cols).T.tolist()
        body_rows.append([instructor,course,semester])

        self.connectivity.num_rows += 1
        
        tobox.insert(tk.END,course)

        self.connectivity.body_cols = np.array(body_rows).T.tolist()

        self.frame_courses.status.insert(tk.END,status)
        self.frame_courses.status.yview(tk.END)

    def test_load(self):

        idx = self.frame_notebook.index(self.frame_notebook.select())

        frameID = "frame"+str(idx)

        f0 = getattr(self,frameID).listbox0.get(tk.ACTIVE)
        f1 = getattr(self,frameID).listbox1.get(tk.ACTIVE)

        if not f0 and not f1:
            status = "Total hours is "+str(0)+".\n"
            self.frame_courses.status.insert(tk.END,status)
            self.frame_courses.status.yview(tk.END)
            return

        s0 = getattr(self,frameID).listbox0.get(0,tk.END)
        s1 = getattr(self,frameID).listbox1.get(0,tk.END)

        annual_load = np.array([np.array(s0+s1)]).T

        idy = np.argwhere(np.array([self.courses.description])==annual_load)[:,1]
            
        hours = self.hours.nparray[idy,:]

        hours_sum = hours.sum(axis=0)

        status = "Total hours is "+str(int(hours_sum[8]))+".\n"

        self.frame_courses.status.insert(tk.END,status)
        self.frame_courses.status.yview(tk.END)

    def add_export(self):

        idx = self.frame_notebook.index(self.frame_notebook.select())

        frameID = "frame"+str(idx)

        f0 = getattr(self,frameID).listbox0.get(tk.ACTIVE)
        f1 = getattr(self,frameID).listbox1.get(tk.ACTIVE)

        if not f0 and not f1:
            status = "No course is selected.\n"
            self.frame_courses.status.insert(tk.END,status)
            self.frame_courses.status.yview(tk.END)
            return

        s0 = getattr(self,frameID).listbox0.get(0,tk.END)
        s1 = getattr(self,frameID).listbox1.get(0,tk.END)

        annual_load = np.array([np.array(s0+s1)]).T

        idy = np.argwhere(np.array([self.courses.description])==annual_load)[:,1]
            
        hour_rows = self.hours.nparray[idy,:]

        if not hasattr(self,"wb"):
            self.wb = openpyxl.Workbook()

        ws = self.wb.create_sheet(self.instructors.fullname[idx])

        names = list(np.array(["code","name"]))
        hours = list(self.hours.definition)

        ws.append(names+hours)

        names_colnum = len(names)
        hours_colnum = len(hours)

        for i in range(names_colnum+1,names_colnum+hours_colnum+1):

            ws.cell(1,i).alignment = Alignment(textRotation=90)

            ws.column_dimensions[get_column_letter(i)].width = 4

        for i,hour_row in enumerate(hour_rows):

            code = self.courses.code[idy[i]]
            name = self.courses.name_AZE[idy[i]]

            ws.append(list(np.array([code,name]))+list(hour_row))

        status = "Saved " + self.instructors.fullname[idx] + "'s schedule to export.\n"
        
        self.frame_courses.status.insert(tk.END,status)
        self.frame_courses.status.yview(tk.END)

    def export_schedule(self):

        if not hasattr(self,"wb"):
            status = "No schedule is added to export.\n"
            self.frame_courses.status.insert(tk.END,status)
            self.frame_courses.status.yview(tk.END)
            return

        exportfilepath = filedialog.asksaveasfilename(
            title = "Export as",
            initialdir = os.getcwd(),
            defaultextension = ".xlsx",
            filetypes=(("Excel Workbook","*.xlsx"),))
        
        self.wb.save(filename=exportfilepath)

        status = "Saved to \""+exportfilepath+"\".\n"
        
        self.frame_courses.status.insert(tk.END,status)
        self.frame_courses.status.yview(tk.END)

        # def check_resize_mode(self,x,y):

        # width = self.frame_notebook.cget('width')
        # # height = self.frame_courses.cget('height')

        # mode = 0

        # if x > width-10: mode |= HORIZONTAL    
        # # if y < 10: mode |= VERTICAL

        # return mode

        # def start_resize(self,event):

        # # print(event.x)
        # self.resize_mode = self.check_resize_mode(event.x,event.y)        
        # # print(self.check_resize_mode(self.frame_courses,event.x,event.y))

        # def resize_frame(self,event):
        # if self.resize_mode:
        #    if self.resize_mode & HORIZONTAL:
        #        self.frame_notebook.config(width=event.x)
        # # if self.resize_mode & VERTICAL:
        # # self.frame_notebook.config(height=event.y)
        # else:
        #    cursor = 'sb_h_double_arrow' if self.check_resize_mode(event.x,event.y) else ''
        #    if cursor != self.cursor:
        #        self.frame_notebook.config(cursor=cursor)
        #        self.cursor = cursor

        # def stop_resize(self,event):
        # self.resize_mode = 0

    def edit_instructors(self):

        self.topEditInstructors = tk.Toplevel()

        self.instructors.draw(self.topEditInstructors,func=self.set_notebook)

        self.topEditInstructors.mainloop()

    def edit_courses(self):

        self.topEditCourses = tk.Toplevel()

        self.courses.draw(self.topEditCourses,func=self.set_courses)

        self.topEditCourses.mainloop()

    def extended_view(self):

        self.topExtendedView = tk.Toplevel()

        self.connectivity.draw(self.topExtendedView)

        self.topExtendedView.mainloop()

    def open(self):
        
        filepath = filedialog.askopenfilename(
            title = "Select a File",
            initialdir = os.getcwd(),
            filetypes = (("BHOS Files","*.bhos"),
                         ("All Files","*")))

        if not filepath: return

        self.instructors = table(filepath,headers=["INSTRUCTORS"],equalsize=False)
        self.courses = table(filepath,headers=["COURSES"],equalsize=False)
        self.connectivity = table(filepath,headers=["CONNECTIVITY"],equalsize=False)

        # self.instructors = table(dm.get_column("links")[0])
        # self.courses = table(dm.get_column("links")[1])
        # self.connectivity = table(dm.get_column("links")[2])

        self.set_notebook()
        self.set_courses()
        # self.set_connections()

    def save(self):

        if not hasattr(self,"outpath"):
            self.saveAs()
            return

        outdir = os.path.split(self.outpath)[0]

        instructors_path = os.path.join(outdir,"instructors.bhos")
        courses_path = os.path.join(outdir,"courses.bhos")
        connectivity_path = os.path.join(outdir,"connectivity.bhos")

        self.instructors.write(instructors_path)
        self.courses.write(courses_path)
        self.connectivity.write(connectivity_path)

        filepaths = [instructors_path,courses_path,connectivity_path]

        with open(self.outpath,'w') as outputfile:
            for filepath in filepaths:
                filename = os.path.split(filepath)[1]
                filename = os.path.splitext(filename)[0]
                outputfile.write(filename.upper())
                outputfile.write("\n")
                with open(filepath) as infile:
                    for line in infile:
                        outputfile.write(line)
                outputfile.write("\n\n\n")
                os.remove(filepath)
            outputfile.write("END")

    def saveAs(self):

        outpath = filedialog.asksaveasfilename(
            title = "Select a File",
            initialdir = os.getcwd(),
            filetypes = (("BHOS Files","*.bhos"),),
            defaultextension = ".bhos")

        if not outpath: return

        outpath = outpath.replace("/","\\")

        self.outpath = outpath

        self.save()
        
if __name__ == "__main__":
    
    window = tk.Tk()
    
    gui = schedule(window)
                                                                  
    # window.geometry("1100x500")

    # Grid.rowconfigure(window, 0, weight=1)
    # Grid.columnconfigure(window, 0, weight=1)

    # frame = Frame(window)

    # Grid.columnconfigure(frame, 0, weight=1)
    # Grid.columnconfigure(frame, 1, weight=1)

    # Grid.rowconfigure(frame, 0, weight=1)

    # listbox1 = Listbox(frame,width=40,height=20)
    # listbox2 = Listbox(frame,width=80,height=20)

    # listbox1.grid(row=0,column=0,sticky=N+E+W+S)
    # listbox2.grid(row=0,column=1,sticky=N+E+W+S)

    # frame.grid(row=0,column=0,sticky=EW)

    # frame.pack(expand=1,fill=BOTH)

    window.mainloop()
