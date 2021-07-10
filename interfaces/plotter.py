"""
The task is about the history matches. When we do changes to input files and run
it, we want to compare two sets of production-history data. It is kinda
having set point between each runs that allow us to select best path to follow.
"""

import datetime
import os

import matplotlib.pyplot as plt

import numpy as np

import openpyxl

import tkinter as tk

from tkinter import filedialog

from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

class plot_production():

    def __init__(self,window):

        self.root = window

        self.initUI()

    def initUI(self):

        self.root.title("BEOC-Subsurface")
        
        self.root.configure(background="white")

        menubar = tk.Menu(self.root)
        
        self.root.config(menu=menubar)

        fileMenu = tk.Menu(menubar,tearoff="Off")

        fileMenu.add_command(label="Open",command=self.set_path)

        menubar.add_cascade(label="File",menu=fileMenu)

        self.frame = tk.Frame(self.root,width=300,height=200)

        tk.Grid.rowconfigure(self.frame,0,weight=1)

        tk.Grid.columnconfigure(self.frame,0,weight=1)
        tk.Grid.columnconfigure(self.frame,1,weight=1)

        self.frame.configure(background="white")

        self.listbox = tk.Listbox(self.frame,width=10,height=30)
        self.listbox.grid(row=0,column=0,sticky=tk.NSEW)

        self.listbox.bind('<Double-1>',self.get_listentry)

        self.button = tk.Button(self.frame,text="Plot Graph",command=lambda: self.get_sheet(self.listbox))
        self.button.grid(row=1,column=0)

        self.figure = plt.Figure()
        
        self.plot = FigureCanvasTkAgg(self.figure,self.frame)

        self.plot_widget = self.plot.get_tk_widget()

        self.plot_widget.grid(row=0,column=1,sticky=tk.NSEW) #width=250,height=30
        
        self.status = tk.Listbox(self.frame,width=250,height=5)
        self.status.grid(row=1,column=1,sticky=tk.EW)

        self.frame.pack(side=tk.LEFT,expand=1,fill=tk.BOTH)
        
    def set_path(self):

        self.filepath = filedialog.askopenfilename(
            title = "Select a File",
            initialdir = os.getcwd(),
            filetypes = (("Excel files","*.xl*"),("All files","*.*")))

        if not self.filepath:
            return

        self.inputfile = self.filepath
        
        self.wb = openpyxl.load_workbook(self.inputfile)

        status = "Imported \""+self.filepath+"\"."

        for sheet in self.wb.sheetnames:
            self.listbox.insert(tk.END,sheet)
        
        self.status.insert(tk.END,status)
        self.status.see(tk.END)

    def get_listentry(self,event):
        
        self.get_sheet(event.widget)

    def get_sheet(self,item):

        if not item.curselection():
            return

        if hasattr(self,"axes"):
            for axis in self.axes:
                axis.remove()

        self.template()

        class data: pass
                
        data.ws = self.wb[self.listbox.get(item.curselection())]
        
        data.date = list(data.ws.iter_cols(min_row=2,max_col=1,values_only=True))[0]
        
        Head = list(data.ws.iter_rows(max_row=1,min_col=2,values_only=True))[0]
        Body = list(data.ws.iter_cols(min_row=2,min_col=2,values_only=True))

        for i,head in enumerate(Head):
            string = head.split(":")[1].split(",")[0].strip().replace(" ","_")
            string = string.replace("(","").replace(")","").replace(".","")
            setattr(data,string,np.array(Body[i]))
        
        try: self.axes[0].plot(data.date,data.Oil_Rate,c='k')
        except: pass
        try: self.axes[0].plot(data.date,data.Oil_Rate_H,'--',c='g')
        except: pass
        try: self.axes[1].plot(data.date,data.Oil_Total/1000,c='k')
        except: pass
        try: self.axes[1].plot(data.date,data.Oil_Total_H/1000,'--',c='g')
        except: pass
        
        try: self.axes[2].plot(data.date,data.Gas_Rate/1000,c='k')
        except: pass
        try: self.axes[2].plot(data.date,data.Gas_Rate_H/1000,'--',c='r')
        except: pass
        try: self.axes[3].plot(data.date,data.Gas_Total/1000000,c='k')
        except: pass
        try: self.axes[3].plot(data.date,data.Gas_Total_H/1000000,'--',c='r')
        except: pass
        
        try: self.axes[4].plot(data.date,data.Water_Rate,c='k')
        except: pass
        try: self.axes[4].plot(data.date,data.Water_Rate_H,'--',c='b')
        except: pass
        try: self.axes[5].plot(data.date,data.Water_Total/1000,c='k')
        except: pass
        try: self.axes[5].plot(data.date,data.Water_Total_H/1000,'--',c='b')
        except: pass
        
        try: self.axes[6].plot(data.date,data.Bottom_Hole_Pressure,c='k')
        except: self.axes[6].plot(data.date,data.Avg_Pressure,c='k')
        try: self.axes[6].plot(data.date,data.Bottom_Hole_Pressure_H,'--',c='m')
        except: pass

        self.figure.set_tight_layout(True)
        
        self.plot.draw()

    def template(self):

        self.axes = []

        self.axes.append(self.figure.add_subplot(221))
        self.axes.append(self.axes[0].twinx())
        self.axes.append(self.figure.add_subplot(222))
        self.axes.append(self.axes[2].twinx())
        self.axes.append(self.figure.add_subplot(223))
        self.axes.append(self.axes[4].twinx())
        self.axes.append(self.figure.add_subplot(224))

        self.axes[0].set_xlabel("Date")
        self.axes[2].set_xlabel("Date")
        self.axes[4].set_xlabel("Date")
        self.axes[6].set_xlabel("Date")

        self.axes[0].set_ylabel("Liquid Rate, sm3/day")
        self.axes[2].set_ylabel("Gas Rate, th. sm3/day")
        self.axes[4].set_ylabel("Liquid Rate, sm3/day")
        self.axes[6].set_ylabel("Pressure, Bars")

        self.axes[1].set_ylabel("Liquid Volume, th. sm3")
        self.axes[3].set_ylabel("Surface Gas Volume, mln. sm3")
        self.axes[5].set_ylabel("Liquid Volume, th. sm3")

        self.axes[0].grid()
        self.axes[2].grid()
        self.axes[4].grid()
        self.axes[6].grid()
        
if __name__ == "__main__":

    window = tk.Tk()

    gui = plot_production(window)

    window.mainloop()
