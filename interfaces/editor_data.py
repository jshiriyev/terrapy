import csv
import datetime

from dateutil.relativedelta import relativedelta

import os
import re
import sys

import matplotlib.pyplot as plt

from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

import numpy as np

import openpyxl

from PyQt5 import QtCore
from PyQt5 import QtGui
from PyQt5 import QtWidgets

import sqlite3

from sqlite3 import Error

import tkinter as tk

from tkinter import ttk
from tkinter import filedialog

class database_manager():

    def __init__(self,dbpath):

        self.dbpath = dbpath

        self.create_connection()

    def create_connection(self):

        self.conn = None

        try:
            self.conn = sqlite3.connect(self.dbpath)
        except Error:
            print(Error)

    def create_table(self,sqlite_table):

        self.sqlite_table = sqlite_table

        try:
            self.cursor = self.conn.cursor()
            self.cursor.execute(self.sqlite_table)
            self.conn.commit()
        except Error:
            print(Error)

    def insert_table(self,sqlite_table_insert,table_row):

        self.sqlite_table_insert = sqlite_table_insert
        self.cursor.execute(self.sqlite_table_insert,table_row)
        self.conn.commit()

class database_interface():

    def __init__(self,window):

        self.root = window

        self.init_interface()

    def init_interface(self):

        self.label_firstname = tk.Label(self.root,text="First Name")
        self.label_firstname.grid(row=0,column=0)

        self.label_lastname = tk.Label(self.root,text="Last Name")
        self.label_lastname.grid(row=1,column=0)

        self.entry_firstname = tk.Entry(self.root)
        self.entry_firstname.grid(row=0,column=1)

        self.entry_lastname = tk.Entry(self.root)
        self.entry_lastname.grid(row=1,column=1)

        self.button_enter = tk.Button(self.root,text="Enter",command=self.save_to_database)
        self.button_enter.grid(row=2,column=1)

    def save_to_database(self):

        db_file = r"C:\Users\Cavid\Documents\bhospy\instructors.db"

        self.conn = sqlite3.connect(db_file)
        # sqlite_table = '''CREATE TABLE SqliteDb_developers (
        #                  id INTEGER PRIMARY KEY,
        #                  name TEXT NOT NULL,
        #                  email text NOT NULL UNIQUE);'''
        self.cursor = self.conn.cursor()
        # self.cursor.execute(sqlite_table)
        # self.conn.commit()
        
        self.cursor.close()
        self.conn.close()

class datafile_manager():

    def __init__(self,filepath,*args,**kwargs):

        self.path = filepath
        self.name = self.path.split("\\")[-1]

        self.extension = os.path.splitext(self.path)[1]

        if self.extension == ".csv":
            self.read_csv(*args)
        elif self.extension == ".xlsx":
            self.read_xlsx(*args)
        else:
            self.read_naked(**kwargs)

    def read_naked(self,skiplines=0,headerline=None):

        self.skiplines = skiplines

        with open(self.path) as structured_text:
            self.running = structured_text.readlines()

        self.body_rows = []

        for linenumber in range(len(self.running)):
            if linenumber>=self.skiplines:
                body_rowline = self.running[linenumber]
                self.body_rows.append(body_rowline.split("\n")[0].split("\t"))

        self.body_rows = np.array(self.body_rows)
        self.num_cols = self.body_rows.shape[1]

        if self.skiplines==0:
            self.header_rows = None
            self.header = ["col_"+str(column_id) for column_id in range(self.num_cols)]
            self._set_columns()
        
        elif self.skiplines!=0:
            if headerline is None:
                linenumber = self.skiplines-1
            elif headerline < self.skiplines:
                linenumber = headerline
            else:
                linenumber = self.skiplines-1
            self.header_rows = self.running[:self.skiplines]
            self.header = self.running[linenumber]
            self.header = self.header.split("\n")[0].split("\t")
            self._set_columns()

        # for linenumber,line in enumerate(self.running):

        #     if rlinenumber>0:

        #         alist = rline.split('\t')

        #         wellname = alist[0]

        #         date_list = alist[1].split('-')

        #         if date_list[1]=='01':
        #             date_newformat = date_list[2]+' JAN '+date_list[0]
        #         elif date_list[1]=='02':
        #             date_newformat = date_list[2]+' FEB '+date_list[0]
        #         elif date_list[1]=='03':
        #             date_newformat = date_list[2]+' MAR '+date_list[0]
        #         elif date_list[1]=='04':
        #             date_newformat = date_list[2]+' APR '+date_list[0]
        #         elif date_list[1]=='05':
        #             date_newformat = date_list[2]+' MAY '+date_list[0]
        #         elif date_list[1]=='06':
        #             date_newformat = date_list[2]+' JUN '+date_list[0]
        #         elif date_list[1]=='07':
        #             date_newformat = date_list[2]+' JUL '+date_list[0]
        #         elif date_list[1]=='08':
        #             date_newformat = date_list[2]+' AUG '+date_list[0]
        #         elif date_list[1]=='09':
        #             date_newformat = date_list[2]+' SEP '+date_list[0]
        #         elif date_list[1]=='10':
        #             date_newformat = date_list[2]+' OCT '+date_list[0]
        #         elif date_list[1]=='11':
        #             date_newformat = date_list[2]+' NOV '+date_list[0]
        #         elif date_list[1]=='12':
        #             date_newformat = date_list[2]+' DEC '+date_list[0]

        #         rdate = np.append(rdate,date_newformat)
                   
        #         days = alist[2]
        #         condensate = alist[3]
        #         gas = alist[4]
        #         water = alist[5]

        #         if rlinenumber+1==len(DFM.content):
        #             copied_line = '\t\''+wellname+'\'\tSTOP\tGRAT\t'+condensate+'\t'+water+'\t'+gas+' /\n'
        #         else:
        #             copied_line = '\t\''+wellname+'\'\tOPEN\tGRAT\t'+condensate+'\t'+water+'\t'+gas+' /\n'

        #         rcopy = np.append(rcopy,copied_line)

    def read_csv(self,*args):

        with open(self.path,"r") as csv_text:
            pass
            # reader = list(csv.reader(csvfile))
            # reader = np.array(reader)
            # num_row, num_col = reader.shape
            # if sheets['dataTypes']=='col':
            #    tuples = tuple(reader.T)
            # elif sheets['dataTypes']=='row':
            #    tuples = tuple(reader)
            # setparent(tuples,*args)
            
    def read_xlsx(self,sheets,*args):
        
        wb = openpyxl.load_workbook(self.path)

        # all_tuples = ()

        # for i,sheetname in enumerate(sheets["names"]):
        #    sheet = wb[sheetname]
        #    if sheets['dataTypes'][i]=='col':
        #        tuples = tuple(sheet.iter_cols(max_col=sheets['num_cols'][i],values_only=True))
        #    elif sheets['dataTypes'][i]=='row':
        #        tuples = tuple(sheet.iter_rows(max_col=sheets['num_cols'][i],values_only=True))
        #    all_tuples = all_tuples+tuples
        # setparent(all_tuples,*args)

    def todatetime(self,attr_name="date",date_format='%m/%d/%Y'):

        dates_string = getattr(self,attr_name)

        dates = []

        for date_string in dates_string:
            dates.append(datetime.datetime.strptime(date_string,date_format))

        setattr(self,attr_name,dates)

    def sortbased(self,attr_name="date"):

        listA = getattr(self,attr_name)

        idx = list(range(len(listA)))

        zipped_lists = zip(listA,idx)
        sorted_pairs = sorted(zipped_lists)

        tuples = zip(*sorted_pairs)

        listA, idx = [ list(tuple) for tuple in  tuples]

        for header in self.header:
            self.toarray(header)
            setattr(self,header,getattr(self,header)[idx])

    def toarray(self,attr_name):

        list_of_string = getattr(self,attr_name)

        try:
            try:
                array_of_string = np.array(list_of_string).astype('int')
            except:
                array_of_string = np.array(list_of_string).astype('float64')
        except:
            array_of_string = np.array(list_of_string)

        setattr(self,attr_name,array_of_string)

    def write_naked(self,outputfile):

        self.outputfile = outputfile
            
        with open(self.outputfile,"w") as empty_file:

            for linenumber,row in enumerate(self.running):
                j = linenumber-self.skiplines
                cond1 = j>0
                cond2 = (self.Date[j]-self.Date[j-1]).days>31
                cond3 = self.Well_Name[j]==self.Well_Name[j-1]
                if cond1 and cond2 and cond3:
                    new_row = []
                    for k,header in enumerate(self.header):
                        if k==0:
                            new_row.append(self.Well_Name[j-1])
                        elif k==1:
                            new_row.append((self.Date[j-1]+relativedelta(months=1)).strftime("%m/%d/%Y"))
                        else:
                            new_row.append(0)
                    empty_file.write("\t".join(np.array(new_row))+"\n")

                empty_file.write(row)

        # with open(self.outputfile,"w") as writtenfile:
        #     for i,date in enumerate(rdate):
        #        skiptoline(rewrittenfile,'DATES',writtenfile)
        #        skiptoline(rewrittenfile,rdate[i],writtenfile)
        #        skiptoline(rewrittenfile,'WCONHIST',writtenfile)
        #        while True:
        #            wline = next(rewrittenfile)
        #            if wline.split('/')[0].strip().split(' ')[0] == '\'BHR_76\'':
        #                notypewritten.write(rcopy[i])
        #                break
        #            else:
        #                notypewritten.write(wline)
        #     while True:
        #        try:
        #            copiedline = next(rewrittenfile)
        #            writtenfile.write(copiedline)
        #        except:
        #            break

    def _skiptoline(self,file_read,keyword,file_written=None):

        while True:
            
            line = next(file_read)

            if file_written is not None:
                file_written.write(line)
            
            if line.split('/')[0].strip() == keyword:
                break

    def _set_columns(self):

        for column_id,header in enumerate(self.header):
            try:
                header = re.sub(r"[^\w]","",header)
                setattr(self,header,self.body_rows[:,column_id].tolist())
                self.header[column_id] = header
            except:
                setattr(self,header,self.body_rows[:,column_id].tolist())

    def _set_rows(self):
        pass
        # array = np.array(atuple[3:]).astype('float64')
        # setattr(self,atuple[1]+'_unit',atuple[2])
        # setattr(self,atuple[1],array)

class Window(QtWidgets.QMainWindow):

    #itemEdited = QtCore.pyqtSignal(item,column)
    
    def __init__(self,filepath):
        
        super(Window,self).__init__()

        self.resize(1200,600)
        
        filepath = "C:\\Users\\Cavid\\Documents\\bhospy\\interfaces\\sample.csv"
        
        self.filepath = filepath
        self.treeView = data_table(parent=self)
        self.treeView.loadCSV(self.filepath)

        self.treeView.autoColumnWidth()

        # self.fileView = FileViewer(parent=self)

        # print(dir(self.treeView))
        
        self.centralwidget = QtWidgets.QWidget(self)
        self.verticalLayout = QtWidgets.QVBoxLayout(self.centralwidget)
        self.verticalLayout.addWidget(self.treeView)
        # self.verticalLayout.addWidget(self.fileView)
        self.setCentralWidget(self.centralwidget)
        
        self.editedFlag = True

        self.show()   

    def closeEvent(self,event):
        
        if self.treeView.editedFlag:
            choice = QtWidgets.QMessageBox.question(self,'Quit',
                "Do you want to save changes?",QtWidgets.QMessageBox.Yes | 
                QtWidgets.QMessageBox.No)

            if choice == QtWidgets.QMessageBox.Yes:
                self.treeView.writeCSV(self.filepath)

        event.accept()

class data_table(QtWidgets.QTreeView):

    def __init__(self,parent=None):
        super(data_table,self).__init__(parent)

        #print(dir(self))

        self.setSortingEnabled(True)

        self.model = QtGui.QStandardItemModel()
        self.setModel(self.model)

        #self.setColumnWidth(0,800)

        self.editedFlag = False

        self.model.itemChanged.connect(self.changeFlag)

    def changeFlag(self):

        self.editedFlag = True

    def mouseDoubleClickEvent(self,event):

        index = self.indexAt(event.pos())
        #print(index.row(),index.column())

        #item = self.model.item(index.row(),index.column())
        #print(item.data(QtCore.Qt.DisplayRole))
        #item.setEditable(True)

        if index.row()>-1 and index.column()>-1:
            self.edit(index)

    def loadCSV(self,filepath):

        csvrunning = list(csv.reader(open(filepath)))

        #self.treeView.setHeaderLabels(csvrunning[0])

        self.model.setHorizontalHeaderLabels(csvrunning[0])

        for line in csvrunning[1:]:
            items = [
                QtGui.QStandardItem(text)
                for text in line
                ]
            self.model.appendRow(items)

    def writeCSV(self,filepath):

        with open(filepath,"w") as fileOutput:
            writer = csv.writer(fileOutput)

            headers = [
                self.model.horizontalHeaderItem(columnNumber).text()
                for columnNumber in range(self.model.columnCount())
            ]

            writer.writerow(headers)

            for rowNumber in range(self.model.rowCount()):
                fields = [
                    self.model.data(
                        self.model.index(rowNumber,columnNumber),
                        QtCore.Qt.DisplayRole
                    )
                    for columnNumber in range(self.model.columnCount())
                ]
                writer.writerow(fields)

    def autoColumnWidth(self):
        
        for columnNumber in range(self.model.columnCount()):
            self.resizeColumnToContents(columnNumber)

class data_graph():

    """
    The task is about the history matches. When we do changes to input files and run
    it, we want to compare two sets of production-history data. It is kinda
    having set point between each runs that allow us to select best path to follow.
    """

    def __init__(self,window):

        self.root = window

        self.init_interface()

    def init_interface(self):
        
        self.root.configure(background="white")

        menubar = tk.Menu(self.root)
        
        self.root.config(menu=menubar)

        fileMenu = tk.Menu(menubar,tearoff="Off")

        fileMenu.add_command(label="Open",command=self.set_path)

        menubar.add_cascade(label="File",menu=fileMenu)

        self.frame_navigator = tk.Frame(self.root,width=300,height=200)
        self.frame_navigator.configure(background="white")

        tk.Grid.rowconfigure(self.frame_navigator,0,weight=1)
        tk.Grid.rowconfigure(self.frame_navigator,2,weight=1)
        tk.Grid.columnconfigure(self.frame_navigator,0,weight=1)

        self.listbox = tk.Listbox(self.frame_navigator,width=10,height=30,exportselection=False)
        self.listbox.grid(row=0,column=0,sticky=tk.NSEW)
        self.listbox.bind('<<ListboxSelect>>',self.get_sheet_data)

        self.label_template = tk.Label(self.frame_navigator,text="Plot Templates")
        self.label_template.grid(row=1,column=0,sticky=tk.EW)

        self.listbox_template = tk.Listbox(self.frame_navigator,exportselection=False)
        self.listbox_template.grid(row=2,column=0,sticky=tk.NSEW)
        self.listbox_template.insert(tk.END,"Production History Match")
        self.listbox_template.bind('<<ListboxSelect>>',self.set_figure_template)

        # self.button = tk.Button(self.frame_navigator,text="Set Plot Template",command=self.set_figure_template)
        # self.button.grid(row=1,column=0,sticky=tk.NSEW)

        self.frame_monitor = tk.Frame(self.root,width=300,height=200)
        self.frame_monitor.configure(background="white")

        tk.Grid.rowconfigure(self.frame_monitor,0,weight=1)
        tk.Grid.columnconfigure(self.frame_monitor,0,weight=1)
        
        self.figure = plt.Figure()
        
        self.plot = FigureCanvasTkAgg(self.figure,self.frame_monitor)
        self.plot_widget = self.plot.get_tk_widget()
        self.plot_widget.grid(row=0,column=0,sticky=tk.NSEW)
        
        self.status = tk.Listbox(self.frame_monitor,width=250,height=5)
        self.status.grid(row=1,column=0,sticky=tk.EW)

        self.frame_navigator.pack(side=tk.LEFT,expand=1,fill=tk.BOTH)
        self.frame_monitor.pack(side=tk.LEFT,expand=1,fill=tk.BOTH)

    def set_path(self):

        self.filepath = filedialog.askopenfilename(
            title = "Select a File",
            initialdir = os.getcwd(),
            filetypes = (("Excel files","*.xl*"),("All files","*.*")))

        if not self.filepath:
            return
        
        self.wb = openpyxl.load_workbook(self.filepath)

        for sheet in self.wb.sheetnames:
            self.listbox.insert(tk.END,sheet)

        status = "Imported \""+self.filepath+"\"."
        self.status.insert(tk.END,status)
        self.status.see(tk.END)

    def set_figure_template(self,event):

        if not self.listbox_template.curselection():
            return
        
        if hasattr(self,"axes"):
            for axis in self.axes:
                self.figure.delaxes(axis)

        self.axes = []

        self.axes.append(self.figure.add_subplot(221))
        self.axes.append(self.axes[0].twinx())
        self.axes.append(self.figure.add_subplot(222))
        self.axes.append(self.axes[2].twinx())
        self.axes.append(self.figure.add_subplot(223))
        self.axes.append(self.axes[4].twinx())
        self.axes.append(self.figure.add_subplot(224))

        self.axes[0].set_xlabel("Date")
        self.axes[2].set_xlabel("Date")
        self.axes[4].set_xlabel("Date")
        self.axes[6].set_xlabel("Date")

        self.axes[0].set_ylabel("Liquid Rate, sm3/day")
        self.axes[2].set_ylabel("Gas Rate, th. sm3/day")
        self.axes[4].set_ylabel("Liquid Rate, sm3/day")
        self.axes[6].set_ylabel("Pressure, Bars")

        self.axes[1].set_ylabel("Liquid Volume, th. sm3")
        self.axes[3].set_ylabel("Surface Gas Volume, mln. sm3")
        self.axes[5].set_ylabel("Liquid Volume, th. sm3")

        self.axes[0].grid()
        self.axes[2].grid()
        self.axes[4].grid()
        self.axes[6].grid()

        self.figure.set_tight_layout(True)

        self.plot.draw()

        status = "Production history match template has been selected."
        self.status.insert(tk.END,status)
        self.status.see(tk.END)

    def get_sheet_data(self,event):

        if not self.listbox.curselection():
            return
        
        class data: pass
                
        data.ws = self.wb[self.listbox.get(self.listbox.curselection())]
        
        data.date = list(data.ws.iter_cols(min_row=2,max_col=1,values_only=True))[0]
        
        Head = list(data.ws.iter_rows(max_row=1,min_col=2,values_only=True))[0]
        Body = list(data.ws.iter_cols(min_row=2,min_col=2,values_only=True))

        for i,head in enumerate(Head):
            string = head.split(":")[1].split(",")[0].strip().replace(" ","_")
            string = string.replace("(","").replace(")","").replace(".","")
            setattr(data,string,np.array(Body[i]))
        
        self.set_figure_data(data)
        
    def set_figure_data(self,data):

        if not hasattr(self,"axes"):
            status = "No template has been selected."    
            self.status.insert(tk.END,status)
            self.status.see(tk.END)
            return

        if hasattr(self,"lines"):
            for line in self.lines:
                line.remove()
                
        self.lines = []
        
        try: self.lines.append(self.axes[0].plot(data.date,data.Oil_Rate,c='k')[0])
        except: pass
        try: self.lines.append(self.axes[0].plot(data.date,data.Oil_Rate_H,'--',c='g')[0])
        except: pass
        try: self.lines.append(self.axes[1].plot(data.date,data.Oil_Total/1000,c='k')[0])
        except: pass
        try: self.lines.append(self.axes[1].plot(data.date,data.Oil_Total_H/1000,'--',c='g')[0])
        except: pass
        
        try: self.lines.append(self.axes[2].plot(data.date,data.Gas_Rate/1000,c='k')[0])
        except: pass
        try: self.lines.append(self.axes[2].plot(data.date,data.Gas_Rate_H/1000,'--',c='r')[0])
        except: pass
        try: self.lines.append(self.axes[3].plot(data.date,data.Gas_Total/1000000,c='k')[0])
        except: pass
        try: self.lines.append(self.axes[3].plot(data.date,data.Gas_Total_H/1000000,'--',c='r')[0])
        except: pass
        
        try: self.lines.append(self.axes[4].plot(data.date,data.Water_Rate,c='k')[0])
        except: pass
        try: self.lines.append(self.axes[4].plot(data.date,data.Water_Rate_H,'--',c='b')[0])
        except: pass
        try: self.lines.append(self.axes[5].plot(data.date,data.Water_Total/1000,c='k')[0])
        except: pass
        try: self.lines.append(self.axes[5].plot(data.date,data.Water_Total_H/1000,'--',c='b')[0])
        except: pass
        
        try: self.lines.append(self.axes[6].plot(data.date,data.Bottom_Hole_Pressure,c='k')[0])
        except: self.lines.append(self.axes[6].plot(data.date,data.Avg_Pressure,c='k')[0])
        try: self.lines.append(self.axes[6].plot(data.date,data.Bottom_Hole_Pressure_H,'--',c='m')[0])
        except: pass

        for axis in self.axes:
            axis.relim()
            axis.autoscale_view()

        self.figure.set_tight_layout(True)

        self.plot.draw()

if __name__ == "__main__":

    """database manager"""

    # dbpath = r"C:\Users\Cavid\Documents\bhospy\interfaces\instructors.db"
    dbpath = ":memory:"
    
    # DB = database_manager(dbpath)

    # instructor_table = """ CREATE TABLE IF NOT EXISTS instructors (
    #                                     id integer PRIMARY KEY,
    #                                     first_name text NOT NULL,
    #                                     last_name text NOT NULL,
    #                                     patronym text NOT NULL,
    #                                     position text NOT NULL,
    #                                     status text NOT NULL,
    #                                     email text NOT NULL UNIQUE,
    #                                     phone integer NOT NULL UNIQUE);"""

    # DB.create_table(instructor_table)

    # instructor_table_insert = """ INSERT INTO instructors(
    #                                     id,
    #                                     first_name,
    #                                     last_name,
    #                                     patronym,
    #                                     position,
    #                                     status,
    #                                     email,
    #                                     phone)
    #                                     VALUES(?,?,?,?,?,?,?,?)"""

    # instructor_7 = (7,"Javid","Shiriyev","Farhad",
    #             "Senior Lecturer","Hour Based Teaching",
    #             "cavid.shiriyev@bhos.edu.az","+994508353992")

    # DB.insert_table(instructor_table_insert,instructor_7)
    # DB.cursor.close()
    # DB.conn.close()

    """database interface"""
    
    # window = tk.Tk()
    # gui = database_constructor(window)
    # window.mainloop()

    """data plot"""
    
    # window = tk.Tk()
    # gui = data_graph(window)
    # window.mainloop()

    """input editor"""

    # app = QtWidgets.QApplication(sys.argv)
    # # window = Window(args.filepath)
    # window = Window(os.curdir)
    # sys.exit(app.exec_())

    """task 0"""
    
    # inputfile0 = 'C:\\Users\\javid.s\\Desktop\\BHR-076.txt'
    # inputfile1 = 'C:\\Users\\javid.s\\Desktop\\BHR_VI_SCHEDULE.INC'

    # outputfile = 'C:\\Users\\javid.s\\Desktop\\BHR_VI_SCHEDULE_V2.INC'

    """input reader"""

    # class reservoir: pass
    # class well: pass
    # class time: pass

    # sheets = {
    #    "num_cols": 5,
    #    "dataTypes": "col"
    #    }

    # setup("sample.csv",sheets,reservoir)

    # print(reservoir.porosity)

    # sheets = {
    #    "names": ["build_up_test","parameters"],
    #    "num_cols": [3,4],
    #    "dataTypes": ["col","row"]
    #    }

    # readxlsx('sample.xlsx',sheets,time,well)
