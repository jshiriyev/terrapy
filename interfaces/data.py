import datetime

import dateutil.parser as parser

import inspect

import os
import re

import numpy as np

import openpyxl

import sqlite3

from sqlite3 import Error as sqlError

class manager():

    special_extensions = [".db",".xlsx"]

    def __init__(self,filepath=None,skiplines=1,headerline=None,comment="--",endline="/",endfile="END",**kwargs):

        if filepath is None:
            return

        self.filepath = filepath

        self.skiplines = skiplines

        if headerline is None:
            self.headerline = skiplines-1
        elif headerline<skiplines:
            self.headerline = headerline
        else:
            self.headerline = skiplines-1

        self.comment = comment
        self.endline = endline
        self.endfile = endfile

        self.filename = os.path.split(self.filepath)[1]
        self.extension = os.path.splitext(self.filepath)[1]

        if any([self.extension==extension for extension in self.special_extensions]):
            self.read_special(**kwargs)
        else:
            self.read_plain()

        self.title = []

        for _ in range(self.skiplines):
            self.title.append(self._running.pop(0))

        num_cols = len(self._running[0])

        if self.skiplines==0:
            self._headers = ["Column #"+str(index) for index in range(num_cols)]
        elif skiplines!=0:
            self._headers = self.title[self.headerline]

        self.headers = self._headers

        nparray = np.array(self._running).T

        self._running = [np.asarray(column) for column in nparray]

        self.running = [np.asarray(column) for column in self._running]

    def read_plain(self):

        # While looping inside the file it does not read lines:
        # - starting with comment phrase, e.g., comment = "--"
        # - after the end of line phrase, e.g., endline = "/"
        # - after the end of file keyword e.g., endfile = "END"

        self._running = []

        with open(self.filepath,"r") as text:

            for line in text:

                line = line.split('\n')[0].strip()

                if self.endline is not None:
                    line = line.strip(self.endline)

                if line=="":
                    continue

                if self.comment is not None:
                    if line[:len(self.comment)] == self.comment:
                        continue

                if self.endfile is not None:
                    if line[:len(self.endfile)] == self.endfile:
                        break

                self._running.append([line])

    def read_special(self,sheetname=None,min_row=1,min_col=1,max_row=None,max_col=None):

        if self.extension == ".xlsx":

            wb = openpyxl.load_workbook(self.filepath,read_only=True)

            lines = wb[sheetname].iter_rows(min_row=min_row,min_col=min_col,
                max_row=max_row,max_col=max_col,values_only=True)

            self._running = [list(line) for line in lines]

            wb._archive.close()
            return

    def set_subheaders(self,header_index=None,header=None,regex=None,regex_builtin="INC_HEADERS",title="SUB-HEADERS"):

        nparray = np.array(self._running[header_index])

        if regex is None and regex_builtin=="INC_HEADERS":
            regex = r'^[A-Z]+$'                         #for strings with only capital letters no digits
        elif regex is None and regex_builtin=="INC_DATES":
            regex = r'^\d{1,2} [A-Z]{3} \d{2}\d{2}? $'   #for strings with [1 or 2 digits][space][3 capital letters][space][2 or 4 digits], e.g. DATES

        vmatch = np.vectorize(lambda x: bool(re.compile(regex).match(x)))

        match_index = vmatch(nparray)

        firstocc = np.argmax(match_index)

        lower = np.where(match_index)[0]
        upper = np.append(lower[1:],nparray.size)

        repeat_count = upper-lower-1

        match_content = nparray[match_index]

        nparray[firstocc:][~match_index[firstocc:]] = np.repeat(match_content,repeat_count)

        self._headers.insert(header_index,title)
        self._running.insert(header_index,np.asarray(nparray))

        for index,column in enumerate(self._running):
            self._running[index] = np.array(self._running[index][firstocc:][~match_index[firstocc:]])

        self.headers = self._headers
        self.running = [np.asarray(column) for column in self._running]

    def texttocolumn(self,header_index=None,header=None,deliminator=None,maxsplit=None):

        if header_index is None:
            header_index = self._header.index(header)

        header_string = self._headers[header_index]
        # header_string = re.sub(deliminator+'+',deliminator,header_string)

        headers = header_string.split(deliminator)

        if maxsplit is not None:
            for index in range(maxsplit):
                if len(headers)<index+1:
                    headers.append("col ##"+string(header_index+index))

        column_to_split = np.asarray(self._running[header_index])

        running = []

        for index,string in enumerate(column_to_split):

            # string = re.sub(deliminator+'+',deliminator,string)
            row = np.char.split(string,deliminator).tolist()

            if maxsplit is not None:
                while len(row)<maxsplit:
                    row.append("")

            running.append(row)

        running = np.array(running,dtype=str).T

        self._headers.pop(header_index)
        self._running.pop(header_index)

        for header,column in zip(headers,running):
            self._headers.insert(header_index,header)
            self._running.insert(header_index,column)
            header_index += 1

        # line = re.sub(r"[^\w]","",line)
        # line = "_"+line if line[0].isnumeric() else line
        # vmatch = np.vectorize(lambda x:bool(re.compile('[Ab]').match(x)))
        
        self.headers = self._headers
        self.running = [np.asarray(column) for column in self._running]

    def columntotext(self,header_new,header_indices=None,headers=None,string=None):

        if header_indices is None:
            header_indices = [self._headers.index(header) for header in headers]

        if string is None:
            string = ("{} "*len(header_indices)).strip()

        vprint = np.vectorize(lambda *args: string.format(*args))

        column_new = [np.asarray(self._running[index]) for index in header_indices]

        column_new = vprint(*column_new)

        self._headers.append(header_new)
        self._running.append(column_new)

        self.headers = self._headers
        self.running = [np.asarray(column) for column in self._running]

    def astype(self,header_index=None,header=None,dtype=None):

        if header_index is None:
            header_index = self._headers.index(header)

        if inspect.isclass(dtype):
            vdate = np.vectorize(lambda x: dtype(x))
        elif type(dtype)==str:
            if type(self._running[header_index][0])==datetime.datetime:
                vdate = np.vectorize(lambda x: x.strftime(dtype))
            elif any([type(self._running[header_index][0])==class_ for class_ in [str,np.str_,np.str]]):
                vdate = np.vectorize(lambda x: parser.parse(x).strftime(dtype))

        self._running[header_index] = vdate(self._running[header_index])

        self.running[header_index] = np.asarray(self._running[header_index])

    def upper(self,header_index=None,header=None):

        if header_index is None:
            header_index = self._headers.index(header)

        self._running[header_index] = np.char.upper(self._running[header_index])

    def set_column(self,column,header_index=None,header_new=None):
        
        if header_index is None:
            if header_new is None:
                header_new = "Col ##"+str(len(self._headers))
            self._headers.append(header_new)
            self.headers = self._headers
            self._running.append(column)
            self.running.append(np.asarray(self._running[-1]))
        else:
            self._running[header_index] = column
            self.running[header_index] = np.asarray(self._running[header_index])

    def set_rows(self,row,row_indices=None):
        
        if row_indices is None:
            for index,column in enumerate(self._running):
                self._running[index] = np.append(column,row[index])
        else:
            for index, _ in enumerate(self._running):
                self._running[index][row_indices] = row[index]

        self.running = [np.asarray(column) for column in self._running]

    def get_rows(self,row_indices):

        if type(row_indices)==int:
            row_indices = [row_indices]

        rows = [[column[index] for column in self._running] for index in row_indices]
        
        return rows

    def get_columns(self,header_indices=None,headers=None,inplace=False):

        if header_indices is None:
            header_indices = [self._headers.index(header) for header in headers]

        if inplace:
            self._headers = [self._headers[index] for index in header_indices]
            self.headers = self._headers
            self._running = [self._running[index] for index in header_indices]
            self.running = [np.asarray(column) for column in self._running]
        else:
            self.headers = [self._headers[index] for index in header_indices]
            self.running = [np.asarray(self._running[index]) for index in header_indices]

    def del_rows(self,row_indices,inplace=False):

        all_rows = np.array([np.arange(self._running[0].size)])

        row_indices = np.array(row_indices).reshape((-1,1))

        comp_mat = all_rows==row_indices

        keep_index = ~np.any(comp_mat,axis=0)

        if inplace:
            self._running = [column[keep_index] for column in self._running]
            self.running = [np.asarray(column) for column in self._running]
        else:
            self.running = [np.asarray(column[keep_index]) for column in self._running]

    def sort(self,header_indices=None,headers=None,reverse=False,inplace=False,returnFlag=False):

        if header_indices is None:
            header_indices = [self.headers.index(header) for header in headers]

        columns = [self._running[index] for index in header_indices]

        columns.reverse()

        sort_index = np.lexsort(columns)

        if reverse:
            sort_index = np.flip(sort_index)

        if inplace:
            self._running = [column[sort_index] for column in self._running]
            self.running = [np.asarray(column) for column in self._running]
        else:
            self.running = [np.asarray(column[sort_index]) for column in self._running]

        if returnFlag:
            return sort_index

    def filter(self,header_index=None,header=None,keywords=None,regex=None,inplace=False):

        if header_index is None:
            header_index = self._headers.index(header)

        if keywords is not None:
            match_array = np.array(keywords).reshape((-1,1))
            match_index = np.any(self._running[header_index]==match_array,axis=0)
        else:
            match_vectr = np.vectorize(lambda x:bool(re.compile(regex).match(x)))
            match_index = match_vectr(self._running[header_index])

        if inplace:
            self._running = [column[match_index] for column in self._running]
            self.running = [np.asarray(column) for column in self._running]
        else:
            self.running = [np.asarray(column[match_index]) for column in self._running]

    def write_to(self,filepath,header_indices=None,headers=None,string=None,**kwargs):

        output_extension = os.path.splitext(filepath)[1]

        if any([output_extension==extension for extension in self.special_extensions]):
            self.write_special(filepath,**kwargs)
            return

        if header_indices is None:
            header_indices = [self._headers.index(header) for header in headers]

        if string is None:
            string = ("{}\t"*len(header_indices))[:-1]+"\n"

        vprint = np.vectorize(lambda *args: string.format(*args))

        columns = [np.asarray(self._running[index]) for index in header_indices]

        with open(filepath,"w",encoding='utf-8') as wfile:
            for line in vprint(*columns):
                wfile.write(line)

    def write_special(self,filepath,**kwargs):

        if extension == ".xlsx":
            wb = openpyxl.Workbook()
            sheet = wb.active
            if sheet_title is not None:
                sheet.title = sheet_title
            for line in running:
                sheet.append(line)
            wb.save(filepath)
            return

class database():

    def __init__(self):

        pass

    def read_database():

        self.conn = None

        try:
            self.conn = sqlite3.connect(self.filepath)
        except Error:
            print(Error)
        return

    def write_database():

        # _set_database_table(self,sqlite_table):

        # instructor_table = """ CREATE TABLE IF NOT EXISTS instructors (
        #                                     id integer PRIMARY KEY,
        #                                     first_name text NOT NULL,
        #                                     last_name text NOT NULL,
        #                                     patronym text NOT NULL,
        #                                     position text NOT NULL,
        #                                     status text NOT NULL,
        #                                     email text NOT NULL UNIQUE,
        #                                     phone integer NOT NULL UNIQUE);"""

        self.sqlite_table = sqlite_table

        try:
            self.cursor = self.conn.cursor()
            self.cursor.execute(self.sqlite_table)
            self.conn.commit()
        except Error:
            print(Error)

        # _insert_database_table(self,sqlite_table_insert,table_row):

        # instructor_table_insert = """ INSERT INTO instructors(
        #                                     id,
        #                                     first_name,
        #                                     last_name,
        #                                     patronym,
        #                                     position,
        #                                     status,
        #                                     email,
        #                                     phone)
        #                                     VALUES(?,?,?,?,?,?,?,?)"""

        self.sqlite_table_insert = sqlite_table_insert
        self.cursor.execute(self.sqlite_table_insert,table_row)
        self.conn.commit()

# function vtkwrite(frac,time,sol)
    
#     % deleteing files in results file
    
#     delete 'results\*.fig'
#     delete 'results\*.vtk'
#     delete 'results\*.out'
    
#     % conversion to field units
    
#     T = time.tau/setup.convFactorDetermine('time');
    
#     Pf = sol.pressure/setup.convFactorDetermine('pressure');
# %   Qf = sol.fracflux/setup.convFactorDetermine('velocity');
    
#     Pw = sol.wellpressure/setup.convFactorDetermine('pressure');
#     Qw = sol.wellflowrate/setup.convFactorDetermine('flowrate');
    
#     % writing time values of well pressure and flowrate
    
#     fid = fopen('results\solution.out','w');
    
#     fprintf(fid,'FRACTURE FLOW ANALYTICAL SOLUTION\r\n');
#     fprintf(fid,'WELL PRESSURE AND FLOW-RATE\r\n');
#     fprintf(fid,'\r\n%-10s\t%-10s\t%-10s\r\n','Time','Pressure','Flow-Rate');
#     fprintf(fid,'%-10s\t%-10s\t%-10s\r\n','[days]','[psi]','[bbl/day]');
    
#     fclose(fid);
    
#     dlmwrite('results\solution.out',[T,Pw',Qw'],'-append',...
#              'delimiter','\t','precision','%-10.3f');
    
#     % writing time values of fracture pressure
    
#     for j = 1:time.numTimeStep
    
#         fid = fopen(['results\fracPressure',num2str(j),'.vtk'],'w');

#         fprintf(fid,'# vtk DataFile Version 1.0\r\n');
#         fprintf(fid,'FRACTURE FLOW ANALYTICAL SOLUTION\r\n');
#         fprintf(fid,'ASCII\r\n');

#         fprintf(fid,'\r\nDATASET UNSTRUCTURED_GRID\r\n');

#         fprintf(fid,'\r\nPOINTS %d FLOAT\r\n',frac.numAnode*2);

#         for i = 1:frac.numAnode
#             fprintf(fid,'%f %f %f\r\n',frac.nodeCoord(i,:));
#         end

#         for i = 1:frac.numAnode
#             fprintf(fid,'%f %f %f\r\n',[frac.nodeCoord(i,1:2),0]);
#         end

#         fprintf(fid,'\r\nCELLS %d %d\r\n',frac.numAfrac,5*frac.numAfrac);

#         for i = 1:frac.numAfrac
#             fprintf(fid,'%d %d %d %d %d\r\n',[4,frac.map(i,:)-1,frac.map(i,:)+frac.numAnode-1]);
#         end

#         fprintf(fid,'\r\nCELL_TYPES %d\r\n',frac.numAfrac);

#         for i = 1:frac.numAfrac
#             fprintf(fid,'%d\r\n',8);
#         end

#         fprintf(fid,'\r\nCELL_DATA %d\r\n',frac.numAfrac);
#         fprintf(fid,'SCALARS pressure float\r\n');
#         fprintf(fid,'LOOKUP_TABLE default\r\n');

#         for i = 1:frac.numAfrac
#             fprintf(fid,'%f\r\n',Pf(i,j));
#         end

#         fclose(fid);
    
#     end

# end

if __name__ == "__main__":

    pass

    """DATA"""

    # """DATABASE"""

    # dbpath = r"C:\Users\Cavid\Documents\bhospy\interfaces\instructors.db"
    
    # DB = database_manager(dbpath)

    # DB.create_table(instructor_table)

    # instructor = (7,"Javid","Shiriyev","Farhad",
    #             "Senior Lecturer","Hour Based Teaching",
    #             "cavid.shiriyev@bhos.edu.az","+994508353992")

    # DB.insert_table(instructor_table_insert,instructor)
    # DB.cursor.close()
    # DB.conn.close()
